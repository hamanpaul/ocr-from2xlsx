from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter

if TYPE_CHECKING:
    from ocr_from2xlsx.form_layout import FormLayout
    from PIL import Image as PILImage

Box = tuple[float, float, float, float]

_TEXT_INSET_X = 2.0
_TEXT_INSET_Y = 1.0
_ENTRY_PAD_X = 6.0
_ENTRY_PAD_Y = 0.0
_FONT_FILES_BY_FAMILY = {
    "times new roman": {
        (False, False): "times.ttf",
        (True, False): "timesbd.ttf",
        (False, True): "timesi.ttf",
        (True, True): "timesbi.ttf",
    },
    "標楷體": {(False, False): "kaiu.ttf"},
    "dfkai-sb": {(False, False): "kaiu.ttf"},
    "新細明體": {(False, False): "mingliu.ttc"},
    "pmingliu": {(False, False): "mingliu.ttc"},
    "微軟正黑體": {(False, False): "msjh.ttc"},
    "microsoft jhenghei": {(False, False): "msjh.ttc"},
}


@dataclass(frozen=True, slots=True)
class SheetGeometry:
    col_x: tuple[float, ...]
    row_y: tuple[float, ...]
    width: int
    height: int
    cell_text: dict[str, str]
    cell_style: dict[str, "CellStyle"]
    span_ref_by_cell: dict[str, str]
    span_anchor_by_ref: dict[str, str]


@dataclass(frozen=True, slots=True)
class CellStyle:
    font_name: str | None
    font_size: float | None
    bold: bool
    italic: bool
    horizontal: str | None
    vertical: str | None


@dataclass(frozen=True, slots=True)
class RenderedLine:
    cell: str
    span_ref: str
    line_index: int
    text: str
    box: Box
    checkbox_box: Box | None


@dataclass(frozen=True, slots=True)
class TemplateRender:
    image: "PILImage.Image"
    span_boxes: dict[str, Box]
    line_boxes: dict[tuple[str, int], RenderedLine]


@dataclass(frozen=True, slots=True)
class _EntrySpec:
    line_cell: str
    line_index: int
    container_ref: str


_TEXT_ENTRY_SPECS: dict[str, _EntrySpec] = {
    "service_date": _EntrySpec("A2", 0, "A2:F2"),
    "medical_record_no": _EntrySpec("B23", 0, "B23"),
    "name": _EntrySpec("B23", 1, "B23"),
    "diagnosis_date": _EntrySpec("A24", 0, "A24:F24"),
}


def _load_font(font_path: str | None, size: int = 12) -> Any:
    from PIL import ImageFont

    if font_path:
        return ImageFont.truetype(font_path, size=size)
    return ImageFont.load_default()


def _resolve_font_path(style: CellStyle, fallback_path: str | None) -> str | None:
    fonts_dir = Path(r"C:\Windows\Fonts")
    family = (style.font_name or "").strip().lower()
    family_files = _FONT_FILES_BY_FAMILY.get(family)
    if family_files:
        candidate_name = family_files.get((style.bold, style.italic)) or family_files.get((False, False))
        if candidate_name:
            candidate = fonts_dir / candidate_name
            if candidate.is_file():
                return str(candidate)
    return fallback_path


def _split_cell(ref: str) -> tuple[int, int]:
    letters = []
    digits = []
    for ch in ref:
        if ch.isalpha():
            letters.append(ch.upper())
        else:
            digits.append(ch)
    col = 0
    for ch in letters:
        col = col * 26 + (ord(ch) - 64)
    row = int("".join(digits))
    return col, row


def cell_box(ref: str, geom: SheetGeometry) -> Box:
    col, row = _split_cell(ref)
    return (geom.col_x[col - 1], geom.row_y[row - 1], geom.col_x[col], geom.row_y[row])


def _span_box(ref: str, geom: SheetGeometry) -> Box:
    if ":" not in ref:
        return cell_box(ref, geom)
    start_ref, end_ref = ref.split(":", 1)
    x0, y0, _, _ = cell_box(start_ref, geom)
    _, _, x1, y1 = cell_box(end_ref, geom)
    return (x0, y0, x1, y1)


def _line_height(draw: Any, font: Any) -> float:
    _, y0, _, y1 = draw.textbbox((0, 0), "中", font=font)
    return float(y1 - y0)


def _render_line(
    draw: Any,
    *,
    cell: str,
    span_ref: str,
    line_index: int,
    text: str,
    origin: tuple[float, float],
    font: Any,
) -> RenderedLine:
    x, y = origin
    draw.text((x, y), text, fill=0, font=font)
    full_box = tuple(map(float, draw.textbbox((x, y), text, font=font)))
    checkbox_box: Box | None = None
    checkbox_index = text.find("□")
    if checkbox_index >= 0:
        prefix = text[:checkbox_index]
        glyph_x = x + (float(draw.textlength(prefix, font=font)) if prefix else 0.0)
        checkbox_box = tuple(map(float, draw.textbbox((glyph_x, y), "□", font=font)))
    return RenderedLine(
        cell=cell,
        span_ref=span_ref,
        line_index=line_index,
        text=text,
        box=full_box,
        checkbox_box=checkbox_box,
    )


def render_sheet_template(geom: SheetGeometry, font_path: str | None = None) -> TemplateRender:
    from PIL import Image, ImageDraw

    image = Image.new("L", (geom.width, geom.height), color=255)
    draw = ImageDraw.Draw(image)

    span_boxes: dict[str, Box] = {}
    line_boxes: dict[tuple[str, int], RenderedLine] = {}
    for cell_ref, text in geom.cell_text.items():
        span_ref = geom.span_ref_by_cell.get(cell_ref, cell_ref)
        span_box = span_boxes.setdefault(span_ref, _span_box(span_ref, geom))
        draw.rectangle(span_box, outline=0, width=1)

        style = geom.cell_style.get(
            cell_ref,
            CellStyle(font_name=None, font_size=12.0, bold=False, italic=False, horizontal="left", vertical="top"),
        )
        size = max(8, int(round(style.font_size or 12.0)))
        font = _load_font(_resolve_font_path(style, font_path), size=size)
        line_step = _line_height(draw, font) + 1.0
        line_texts = text.splitlines() or [text]
        line_widths = [float(draw.textbbox((0, 0), line_text, font=font)[2]) for line_text in line_texts]
        total_height = max(0.0, line_step * len(line_texts) - 1.0)

        x0, y0, x1, y1 = span_box
        if style.vertical == "center":
            base_y = max(y0 + _TEXT_INSET_Y, y0 + ((y1 - y0) - total_height) / 2.0)
        else:
            base_y = y0 + _TEXT_INSET_Y

        for line_index, line_text in enumerate(line_texts):
            line_width = line_widths[line_index]
            if style.horizontal == "center":
                base_x = max(x0 + _TEXT_INSET_X, x0 + ((x1 - x0) - line_width) / 2.0)
            elif style.horizontal == "right":
                base_x = max(x0 + _TEXT_INSET_X, x1 - line_width - _TEXT_INSET_X)
            else:
                base_x = x0 + _TEXT_INSET_X
            line = _render_line(
                draw,
                cell=cell_ref,
                span_ref=span_ref,
                line_index=line_index,
                text=line_text,
                origin=(base_x, base_y + line_index * line_step),
                font=font,
            )
            line_boxes[(cell_ref, line_index)] = line

    return TemplateRender(image=image, span_boxes=span_boxes, line_boxes=line_boxes)


def draw_base_form(layout: "FormLayout", geom: SheetGeometry, font_path: str | None = None) -> Any:
    return render_sheet_template(geom, font_path=font_path).image


def _line_entry_box(rendered: TemplateRender, spec: _EntrySpec, geom: SheetGeometry) -> Box:
    line = rendered.line_boxes[(spec.line_cell, spec.line_index)]
    container = rendered.span_boxes.get(spec.container_ref, _span_box(spec.container_ref, geom))
    x0 = max(line.box[2] + _ENTRY_PAD_X, container[0] + _TEXT_INSET_X)
    y0 = max(line.box[1] - _ENTRY_PAD_Y, container[1] + _TEXT_INSET_Y)
    x1 = container[2] - _TEXT_INSET_X
    y1 = min(line.box[3] + _ENTRY_PAD_Y, container[3] - _TEXT_INSET_Y)
    return (x0, y0, x1, y1)


def _synthetic_checkbox_box(line: RenderedLine, option_cell_box: Box) -> Box:
    line_height = line.box[3] - line.box[1]
    size = min(line_height, option_cell_box[3] - option_cell_box[1] - 2.0)
    right = max(option_cell_box[0] + _TEXT_INSET_X + size, line.box[0] - 2.0)
    left = max(option_cell_box[0] + _TEXT_INSET_X, right - size)
    top = max(option_cell_box[1] + _TEXT_INSET_Y, line.box[1] + (line_height - size) / 2.0)
    bottom = min(option_cell_box[3] - _TEXT_INSET_Y, top + size)
    return (left, top, min(option_cell_box[2] - _TEXT_INSET_X, left + size), bottom)


def text_entry_box(
    layout: "FormLayout",
    geom: SheetGeometry,
    field_key: str,
    rendered: TemplateRender | None = None,
) -> Box:
    del layout
    spec = _TEXT_ENTRY_SPECS[field_key]
    template = rendered or render_sheet_template(geom)
    return _line_entry_box(template, spec, geom)


def option_mark_box(
    layout: "FormLayout",
    geom: SheetGeometry,
    field_key: str,
    option_code: str,
    rendered: TemplateRender | None = None,
) -> Box:
    template = rendered or render_sheet_template(geom)
    option = layout.options_by_code(field_key)[option_code]
    normalized_label = option.label.replace(" ", "").replace("\n", "")
    option_cell_box = cell_box(option.cell, geom)
    candidate_lines = [line for (cell, _), line in template.line_boxes.items() if cell == option.cell]
    for line in candidate_lines:
        if normalized_label and normalized_label in line.text.replace("□", "").replace(" ", "").replace("\n", ""):
            return line.checkbox_box or _synthetic_checkbox_box(line, option_cell_box)
    if candidate_lines:
        return candidate_lines[0].checkbox_box or _synthetic_checkbox_box(candidate_lines[0], option_cell_box)
    raise KeyError(f"No checkbox line rendered for {field_key}.{option_code}")


def sheet_geometry(xlsx_path: str | Path, sheet_name: str = "服務紀錄表") -> SheetGeometry:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet_name]

    col_x = [0.0]
    for col in range(1, ws.max_column + 1):
        width = ws.column_dimensions[get_column_letter(col)].width or 8.43
        col_x.append(col_x[-1] + width * 7.0)

    row_y = [0.0]
    for row in range(1, ws.max_row + 1):
        height = ws.row_dimensions[row].height or 15.0
        row_y.append(row_y[-1] + height)

    span_ref_by_cell: dict[str, str] = {}
    span_anchor_by_ref: dict[str, str] = {}
    for merged_range in ws.merged_cells.ranges:
        span_ref = str(merged_range)
        anchor = ws.cell(merged_range.min_row, merged_range.min_col).coordinate
        span_anchor_by_ref[span_ref] = anchor
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                span_ref_by_cell[ws.cell(row, col).coordinate] = span_ref

    cell_text: dict[str, str] = {}
    cell_style: dict[str, CellStyle] = {}
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            value = cell.value
            if value in (None, ""):
                continue
            ref = cell.coordinate
            span_ref = span_ref_by_cell.get(ref)
            if span_ref and span_anchor_by_ref[span_ref] != ref:
                continue
            cell_text[ref] = str(value)
            cell_style[ref] = CellStyle(
                font_name=cell.font.name,
                font_size=cell.font.size,
                bold=bool(cell.font.bold),
                italic=bool(cell.font.italic),
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical,
            )
            span_ref_by_cell.setdefault(ref, ref)
            span_anchor_by_ref.setdefault(ref, ref)

    return SheetGeometry(
        col_x=tuple(col_x),
        row_y=tuple(row_y),
        width=int(col_x[-1]),
        height=int(row_y[-1]),
        cell_text=cell_text,
        cell_style=cell_style,
        span_ref_by_cell=span_ref_by_cell,
        span_anchor_by_ref=span_anchor_by_ref,
    )
