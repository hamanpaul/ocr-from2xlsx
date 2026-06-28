from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from ocr_from2xlsx.constants import (
    AGE_GROUP_LABELS,
    BASIC_COLUMN_BY_FIELD,
    CANCER_LABELS,
    CHANNEL_LABELS,
    DISEASE_STATUS_LABELS,
    GENDER_LABELS,
    IDENTITY_LABELS,
    NATIONALITY_LABELS,
    SOURCE_LABELS,
    WORKBOOK_SHEET,
)
from ocr_from2xlsx.domain import Record
from ocr_from2xlsx.form_layout import service_record_layout

# Column-header prefix for each service group in 個案總表. The per-option NUMBER (its 1-based
# position in the curated form layout) is the column suffix — e.g. symptom option 4
# (fatigue_strength / 4.疲憊與體力) → 諮詢-症狀與副作用照護4. Built from form_layout so the
# written value (the real "N.中文" label) and the target column always match the form
# (#service-write-mapping; previously a 6-entry LABEL_BY_CODE wrote raw English codes to the
# wrong column for every other service option).
_SERVICE_PREFIX_BY_PATH = {
    "services.consultation.health_medical": "諮詢-健康與醫療系統",
    "services.consultation.symptom_side_effect": "諮詢-症狀與副作用照護",
    "services.consultation.nutrition_diet": "諮詢-營養與飲食",
    "services.consultation.psychosocial_emotion": "諮詢-社會心理情緒",
    "services.consultation.financial_social": "諮詢-經濟與社會資源",
    "services.consultation.care_support": "諮詢-照顧與支持",
    "services.supplies": "提供實體用品及設備",
    "services.internal_referrals": "轉介或連結院內資源",
    "services.external_referrals": "轉介或連結院外資源",
    "services.referral_outcomes": "轉介或連結資源成果",
}


def _build_service_index() -> "dict[str, tuple[str, dict[str, tuple[int, str]]]]":
    """record_path → (column prefix, {code: (number, label)}) from the curated form layout."""
    layout = service_record_layout()
    index: dict[str, tuple[str, dict[str, tuple[int, str]]]] = {}
    for field in layout.iter_fields():
        prefix = _SERVICE_PREFIX_BY_PATH.get(field.record_path or "")
        if prefix is None:
            continue
        index[field.record_path] = (
            prefix,
            {opt.code: (i, opt.label) for i, opt in enumerate(field.options, start=1)},
        )
    return index


_SERVICE_INDEX = _build_service_index()


def _build_service_label_to_code() -> "dict[tuple[str, str], str]":
    """Reverse of _SERVICE_INDEX: (column prefix, written label) → option code, so a re-opened
    workbook turns each written service label back into its canonical code. This keeps the
    duplicate-detection summary identical to ``Services.summary()`` for EVERY option, not just
    a hand-picked few (#service-write-mapping)."""
    rev: dict[tuple[str, str], str] = {}
    for prefix, by_code in _SERVICE_INDEX.values():
        for code, (_number, label) in by_code.items():
            rev[(prefix, label)] = code
    return rev


_SERVICE_LABEL_TO_CODE = _build_service_label_to_code()

SUMMARY_PREFIX_BY_HEADER_PREFIX = {
    "諮詢-健康與醫療系統": "health_medical",
    "諮詢-症狀與副作用照護": "symptom_side_effect",
    "諮詢-營養與飲食": "nutrition_diet",
    "諮詢-社會心理情緒": "psychosocial_emotion",
    "諮詢-經濟與社會資源": "financial_social",
    "諮詢-照顧與支持": "care_support",
    "提供實體用品及設備": "supplies",
    "轉介或連結院內資源": "internal",
    "轉介或連結院外資源": "external",
    "轉介或連結資源成果": "outcomes",
}

MAX_CONSECUTIVE_EMPTY_DUPLICATE_ROWS = 25


def _normalize_duplicate_value(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_service_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _normalize_duplicate_value(value)


class WorkbookWriter:
    def __init__(self, working_path: Path | str) -> None:
        self.working_path = Path(working_path)
        suffix = self.working_path.suffix.lower()
        if suffix in {".xlsm", ".xltm"}:
            self.workbook = load_workbook(self.working_path, keep_vba=True)
        else:
            self.workbook = load_workbook(self.working_path)
        self._materialize_embedded_images()
        if WORKBOOK_SHEET not in self.workbook.sheetnames:
            raise ValueError(f"Missing sheet: {WORKBOOK_SHEET}")
        self.sheet = self.workbook[WORKBOOK_SHEET]
        self.header_map = self._build_header_map(self.sheet)

    def _materialize_embedded_images(self) -> None:
        """Cache every embedded image's encoded bytes right after load and make ``_data()``
        return them on every save, so saving never re-reads the image lazily from the source
        archive. openpyxl's lazy re-read intermittently raised ``ValueError: I/O operation on
        closed file`` mid-save (GC-closed zip) and left a CORRUPT workbook — fatal for the
        official template, which carries a logo on every sheet (#xlsx-image-save).

        We cache the *bytes* (not a BytesIO): a single stream gets consumed/closed by the first
        save, so the SECOND write to the same workbook would fail — exactly what a multi-record
        import session does. Returning cached bytes survives unlimited saves. Best-effort: any
        per-image failure leaves that image untouched."""
        for worksheet in self.workbook.worksheets:
            for image in list(getattr(worksheet, "_images", [])):
                try:
                    data = image._data()  # encode now, while the archive is still readable
                except Exception:
                    continue
                try:
                    image._data = lambda _cached=data: _cached
                except Exception:
                    pass

    @classmethod
    def create_from_template(cls, template_path: Path | str, working_path: Path | str) -> "WorkbookWriter":
        template_path = Path(template_path)
        working_path = Path(working_path)
        macro_suffixes = {".xlsm", ".xltm"}
        template_suffix = template_path.suffix.lower()
        working_suffix = working_path.suffix.lower()
        if template_suffix in macro_suffixes and working_suffix not in macro_suffixes:
            raise ValueError("Macro-enabled template requires macro-enabled output path.")
        working_path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(template_path, working_path)
        return cls(working_path)

    def write_record(self, record: Record, row: int | None = None) -> int:
        if row is None:
            row = self._next_empty_row()
        else:
            self._clear_row(row)
        # Tolerate missing/optional values: the relaxed (human-confirmed) GUI write only
        # requires a name, so a blank/invalid date or an unset identity/gender must land as
        # an empty cell rather than crash. The strict import path blocks those upstream, so
        # this only ever softens the relaxed path — valid values map identically.
        try:
            service_month = record.service_month_label()
        except ValueError:
            service_month = ""
        self._set(row, BASIC_COLUMN_BY_FIELD["service_month"], service_month)
        self._set(row, BASIC_COLUMN_BY_FIELD["service_date"], record.service_date)
        self._set(row, BASIC_COLUMN_BY_FIELD["identity"], IDENTITY_LABELS.get(record.identity, ""))
        self._set(row, BASIC_COLUMN_BY_FIELD["name"], record.name)
        self._set(row, BASIC_COLUMN_BY_FIELD["medical_record_no"], record.medical_record_no)
        self._set(row, BASIC_COLUMN_BY_FIELD["gender"], GENDER_LABELS.get(record.gender, ""))

        if record.identity == "patient":
            fields = record.patient_fields
            self._set(row, BASIC_COLUMN_BY_FIELD["nationality"], NATIONALITY_LABELS.get(fields.nationality or ""))
            self._set(row, BASIC_COLUMN_BY_FIELD["age_group"], AGE_GROUP_LABELS.get(fields.age_group or ""))
            self._set(row, BASIC_COLUMN_BY_FIELD["channel"], CHANNEL_LABELS.get(fields.channel or ""))
            self._set(row, BASIC_COLUMN_BY_FIELD["disease_status"], DISEASE_STATUS_LABELS.get(fields.disease_status or ""))
            self._set(row, BASIC_COLUMN_BY_FIELD["source"], SOURCE_LABELS.get(fields.source or ""))
            newly_diagnosed = None
            if fields.newly_diagnosed_within_year is True:
                newly_diagnosed = "是"
            elif fields.newly_diagnosed_within_year is False:
                newly_diagnosed = "否"
            self._set(row, BASIC_COLUMN_BY_FIELD["newly_diagnosed_within_year"], newly_diagnosed)
            for index, cancer in enumerate(fields.cancers[:3], start=1):
                self._set(row, f"癌別{index}\n(病人才填)", CANCER_LABELS.get(cancer, cancer))

        discharge_followup = None
        if record.discharge_followup is True:
            discharge_followup = "是"
        elif record.discharge_followup is False:
            discharge_followup = "否"
        self._set(row, BASIC_COLUMN_BY_FIELD["discharge_followup"], discharge_followup)

        self._write_services(row, record)
        return row

    def save(self) -> None:
        self.workbook.calculation.fullCalcOnLoad = True
        self.workbook.calculation.forceFullCalc = True
        self.workbook.save(self.working_path)

    def close(self) -> None:
        vba_archive = getattr(self.workbook, "vba_archive", None)
        if vba_archive is not None:
            vba_closer = getattr(vba_archive, "close", None)
            if callable(vba_closer):
                vba_closer()
        closer = getattr(self.workbook, "close", None)
        if callable(closer):
            closer()

    def __enter__(self) -> "WorkbookWriter":
        return self

    def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
        self.close()

    def existing_duplicate_keys(self) -> set[tuple[str, str, str, str]]:
        keys: set[tuple[str, str, str, str]] = set()
        service_date_col = self.header_map.get(BASIC_COLUMN_BY_FIELD["service_date"])
        name_col = self.header_map.get(BASIC_COLUMN_BY_FIELD["name"])
        id_col = self.header_map.get(BASIC_COLUMN_BY_FIELD["medical_record_no"])
        if not service_date_col or not name_col or not id_col:
            return keys
        consecutive_empty_rows = 0
        for row in range(2, self.sheet.max_row + 1):
            service_date = _normalize_service_date(self.sheet.cell(row=row, column=service_date_col).value)
            name = _normalize_duplicate_value(self.sheet.cell(row=row, column=name_col).value)
            medical_id = _normalize_duplicate_value(self.sheet.cell(row=row, column=id_col).value)
            if not service_date and not name and not medical_id:
                consecutive_empty_rows += 1
                if consecutive_empty_rows >= MAX_CONSECUTIVE_EMPTY_DUPLICATE_ROWS:
                    break
                continue
            consecutive_empty_rows = 0
            if not service_date or not name or not medical_id:
                continue
            keys.add((service_date, name, medical_id, self._service_summary_from_row(row)))
        return keys

    def _write_services(self, row: int, record: Record) -> None:
        for category, codes in record.services.consultation.items():
            self._write_service_group(row, f"services.consultation.{category}", codes)
        self._write_service_group(row, "services.supplies", record.services.supplies)
        self._write_service_group(row, "services.internal_referrals", record.services.internal_referrals)
        self._write_service_group(row, "services.external_referrals", record.services.external_referrals)
        self._write_service_group(row, "services.referral_outcomes", record.services.referral_outcomes)

    def _write_service_group(self, row: int, record_path: str, codes) -> None:
        """Write each selected code to its OWN numbered column (prefix+number) with the real
        form label. The number is the option's position in the form layout, which equals the
        個案總表 column number — so fatigue_strength lands in 諮詢-症狀與副作用照護4 as
        "4.疲憊與體力", never the raw English code in the first free column
        (#service-write-mapping)."""
        entry = _SERVICE_INDEX.get(record_path)
        if entry is None:
            return
        prefix, by_code = entry
        for code in codes:
            num_label = by_code.get(code)
            if num_label is None:
                # Unknown code (data drift): keep the old safety net so nothing is dropped.
                header = self._first_empty_service_header(row, prefix, code)
                self._set(row, header, code)
                continue
            number, label = num_label
            header = f"{prefix}{number}"
            if header not in self.header_map:
                header = self._first_empty_service_header(row, prefix, code)
            self._set(row, header, label)

    def _first_empty_service_header(self, row: int, prefix: str, code: str) -> str:
        candidates = [header for header in self.header_map if header.startswith(prefix)]
        if not candidates:
            raise ValueError(f"Missing workbook service columns for {prefix}")
        for header in candidates:
            if self._service_cell_empty(row, header):
                return header
        raise ValueError(f"No available workbook column for service {code}")

    def _service_cell_empty(self, row: int, header: str) -> bool:
        column = self.header_map[header]
        return self.sheet.cell(row=row, column=column).value in (None, "")

    def _service_summary_from_row(self, row: int) -> str:
        parts: list[str] = []
        for header, column in self.header_map.items():
            value = self.sheet.cell(row=row, column=column).value
            if value in (None, ""):
                continue
            value_str = str(value)
            for prefix, summary_prefix in SUMMARY_PREFIX_BY_HEADER_PREFIX.items():
                if header.startswith(prefix):
                    # Reverse the written label back to its code so the summary is code-based
                    # and matches Services.summary(); fall back to the raw value on data drift.
                    code = _SERVICE_LABEL_TO_CODE.get((prefix, value_str), value_str)
                    parts.append(f"{summary_prefix}:{code}")
                    break
        return "|".join(sorted(parts))

    def _build_header_map(self, sheet: Worksheet) -> dict[str, int]:
        header_map: dict[str, int] = {}
        for cell in sheet[1]:
            if cell.value:
                header_map[str(cell.value)] = cell.column
        required_headers = list(BASIC_COLUMN_BY_FIELD.values()) + [
            "癌別1\n(病人才填)",
            "癌別2\n(病人才填)",
            "癌別3\n(病人才填)",
        ]
        missing = [header for header in required_headers if header not in header_map]
        if missing:
            raise ValueError(f"Missing required headers: {', '.join(missing)}")
        return header_map

    def _next_empty_row(self) -> int:
        name_col = self.header_map[BASIC_COLUMN_BY_FIELD["name"]]
        for row in range(2, self.sheet.max_row + 2):
            if self.sheet.cell(row=row, column=name_col).value in (None, ""):
                return row
        return self.sheet.max_row + 1

    def _clear_row(self, row: int) -> None:
        # Blank every mapped column in the row so a re-write (overwrite) leaves no
        # stale value from the record that previously occupied it. Assign .value
        # directly: ws.cell(row, col, value=None) is a getter (None means "don't set").
        for column in self.header_map.values():
            self.sheet.cell(row=row, column=column).value = None

    def _set(self, row: int, header: str, value: object) -> None:
        column = self.header_map[header]
        self.sheet.cell(row=row, column=column, value=value)

    def _set_if_present(self, row: int, header: str, value: object) -> None:
        column = self.header_map.get(header)
        if column:
            self.sheet.cell(row=row, column=column, value=value)
