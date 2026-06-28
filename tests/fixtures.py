from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill

HEADERS = [
    "序",
    "服務月份",
    "身分",
    "服務日期",
    "姓名",
    "ID",
    "生日",
    "性別",
    "是否曾經今年服務過",
    "國籍\n(病人才填)",
    "年齡\n(病人才填)",
    "管道\n(病人才填)",
    "疾病狀態\n(病人才填)",
    "來源\n(病人才填)",
    "癌別1\n(病人才填)",
    "癌別2\n(病人才填)",
    "癌別3\n(病人才填)",
    "一年內新診斷(病人才填)",
    "諮詢-健康與醫療系統1",
    "諮詢-營養與飲食1",
    "提供實體用品及設備1",
    "轉介或連結院內資源3",
    "轉介或連結院外資源9",
    "轉介或連結資源成果4",
    "出院後關懷",
    "諮詢-症狀與副作用照護1",
    "諮詢-社會心理情緒1",
    "諮詢-經濟與社會資源1",
    "諮詢-照顧與支持1",
]


def create_workbook_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "個案總表"
    wb.create_sheet("一月")
    for column, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=column, value=header)
        cell.fill = PatternFill("solid", fgColor="FFD966")
    for row in range(2, 7):
        ws.cell(row=row, column=1, value=row - 1)
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    wb["一月"]["A1"] = "=SUM(個案總表!A2:A6)"
    wb.save(path)


# Full 個案總表 service-column layout (prefix, count) mirroring the real workbook, so a test
# can exercise EVERY service option landing in its own numbered column (#service-write-mapping).
SERVICE_GROUPS = [
    ("諮詢-健康與醫療系統", 8),
    ("諮詢-症狀與副作用照護", 7),
    ("諮詢-營養與飲食", 4),
    ("諮詢-社會心理情緒", 6),
    ("諮詢-經濟與社會資源", 8),
    ("諮詢-照顧與支持", 6),
    ("提供實體用品及設備", 4),
    ("轉介或連結院內資源", 10),
    ("轉介或連結院外資源", 10),
    ("轉介或連結資源成果", 4),
]

_FULL_BASIC_HEADERS = [
    "序", "服務月份", "身分", "服務日期", "姓名", "ID", "生日", "性別", "是否曾經今年服務過",
    "國籍\n(病人才填)", "年齡\n(病人才填)", "管道\n(病人才填)", "疾病狀態\n(病人才填)",
    "來源\n(病人才填)", "癌別1\n(病人才填)", "癌別2\n(病人才填)", "癌別3\n(病人才填)",
    "一年內新診斷(病人才填)", "出院後關懷",
]


def full_workbook_headers() -> list[str]:
    headers = list(_FULL_BASIC_HEADERS)
    for prefix, count in SERVICE_GROUPS:
        headers.extend(f"{prefix}{i}" for i in range(1, count + 1))
    return headers


def create_full_workbook_template(path: Path) -> None:
    """Template with the COMPLETE service-column set (every 諮詢-…N / 轉介…N column), unlike the
    minimal create_workbook_template — for exercising the full per-column write mapping."""
    wb = Workbook()
    ws = wb.active
    ws.title = "個案總表"
    wb.create_sheet("一月")
    for column, header in enumerate(full_workbook_headers(), start=1):
        ws.cell(row=1, column=column, value=header)
    wb.save(path)
