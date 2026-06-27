from __future__ import annotations

import re
from datetime import datetime
from pathlib import Path

import pytest

from openpyxl import load_workbook

import ocr_from2xlsx.workbook as workbook_module
from ocr_from2xlsx.constants import BASIC_COLUMN_BY_FIELD
from ocr_from2xlsx.domain import Services
from ocr_from2xlsx.workbook import WorkbookWriter
from tests.fixtures import create_workbook_template
from tests.test_json_io import make_record


def _column_for_header(sheet, header: str) -> int:
    for cell in sheet[1]:
        if cell.value == header:
            return cell.column
    raise AssertionError(f"Missing header in fixture: {header}")


def test_writer_copies_template_and_writes_record(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    writer = WorkbookWriter.create_from_template(template, working)
    row_number = writer.write_record(make_record())
    writer.save()
    writer.close()

    wb = load_workbook(working)
    ws = wb["個案總表"]
    assert row_number == 2
    assert ws["B2"].value == "3月"
    assert ws["C2"].value == "病人"
    assert ws["D2"].value == "2026-03-15"
    assert ws["E2"].value == "王小明"
    assert ws["F2"].value == "A123456"
    assert ws["G2"].value is None
    assert ws["H2"].value == "女性"
    assert ws["I2"].value is None
    assert ws["J2"].value == "本國籍"
    assert ws["K2"].value == "51-60歲"
    assert ws["O2"].value == "8.乳癌"
    assert ws["R2"].value == "是"
    consult_col = _column_for_header(ws, "諮詢-健康與醫療系統1")
    supplies_col = _column_for_header(ws, "提供實體用品及設備1")
    assert ws.cell(row=2, column=consult_col).value == "1.癌症篩檢與預防"
    assert ws.cell(row=2, column=supplies_col).value == "1.假髮/頭巾/毛帽用品"
    assert wb["一月"]["A1"].value == "=SUM(個案總表!A2:A6)"
    wb.close()


def test_writer_preserves_style_and_column_width(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    before = load_workbook(template)
    before_fill = before["個案總表"]["B1"].fill.fgColor.rgb
    before_width = before["個案總表"].column_dimensions["B"].width
    before.close()

    writer = WorkbookWriter.create_from_template(template, working)
    writer.write_record(make_record())
    writer.save()
    writer.close()

    after = load_workbook(working)
    assert after["個案總表"]["B1"].fill.fgColor.rgb == before_fill
    assert after["個案總表"].column_dimensions["B"].width == before_width
    after.close()


@pytest.mark.parametrize("suffix", [".xlsm", ".xltm"])
def test_workbook_writer_sets_keep_vba_for_macro_templates(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, suffix: str
) -> None:
    template = tmp_path / f"template{suffix}"
    create_workbook_template(template)
    captured: dict[str, object] = {}
    real_load_workbook = load_workbook

    def fake_load_workbook(path: Path, **kwargs: object):
        captured.clear()
        captured.update(kwargs)
        return real_load_workbook(path, **kwargs)

    monkeypatch.setattr(workbook_module, "load_workbook", fake_load_workbook)

    writer = WorkbookWriter(template)
    writer.close()

    assert captured.get("keep_vba") is True


def test_workbook_writer_does_not_set_keep_vba_for_xlsx(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    template = tmp_path / "template.xlsx"
    create_workbook_template(template)
    captured: dict[str, object] = {}
    real_load_workbook = load_workbook

    def fake_load_workbook(path: Path, **kwargs: object):
        captured.clear()
        captured.update(kwargs)
        return real_load_workbook(path, **kwargs)

    monkeypatch.setattr(workbook_module, "load_workbook", fake_load_workbook)

    writer = WorkbookWriter(template)
    writer.close()

    assert "keep_vba" not in captured


@pytest.mark.parametrize("suffix", [".xlsm", ".xltm"])
def test_create_from_template_rejects_macro_template_to_xlsx(
    tmp_path: Path, suffix: str
) -> None:
    template = tmp_path / f"template{suffix}"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    with pytest.raises(ValueError, match=re.escape("macro-enabled")):
        WorkbookWriter.create_from_template(template, working)


def test_existing_duplicate_keys_include_service_summary(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    record = make_record()
    writer = WorkbookWriter.create_from_template(template, working)
    writer.write_record(record)
    writer.save()
    writer.close()

    reopened = WorkbookWriter(working)

    assert record.duplicate_key() in reopened.existing_duplicate_keys()
    reopened.close()


def test_existing_duplicate_keys_stops_after_consecutive_empty_rows(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    record = make_record()
    writer = WorkbookWriter.create_from_template(template, working)
    writer.write_record(record)
    writer.save()
    writer.close()

    wb = load_workbook(working)
    ws = wb["個案總表"]
    ws.cell(row=500, column=1, value="inflate")
    wb.save(working)
    wb.close()

    reopened = WorkbookWriter(working)
    max_row_seen = 0
    real_cell = reopened.sheet.cell

    def tracking_cell(*args: object, **kwargs: object):
        nonlocal max_row_seen
        row = kwargs.get("row")
        if row is None and args:
            row = args[0]
        if isinstance(row, int) and row > max_row_seen:
            max_row_seen = row
        return real_cell(*args, **kwargs)

    monkeypatch.setattr(reopened.sheet, "cell", tracking_cell)

    keys = reopened.existing_duplicate_keys()
    reopened.close()

    assert record.duplicate_key() in keys
    assert max_row_seen <= 2 + workbook_module.MAX_CONSECUTIVE_EMPTY_DUPLICATE_ROWS


def test_writer_maps_numbered_services_to_numbered_columns(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    record = make_record()
    record.services = Services(
        consultation={"health_medical": ["screening_prevention"]},
        supplies=["wig_hat"],
        internal_referrals=["social_welfare"],
        external_referrals=["care_information"],
        referral_outcomes=["received_service_help"],
    )

    writer = WorkbookWriter.create_from_template(template, working)
    writer.write_record(record)
    writer.save()
    writer.close()

    wb = load_workbook(working)
    ws = wb["個案總表"]
    internal_col = _column_for_header(ws, "轉介或連結院內資源3")
    external_col = _column_for_header(ws, "轉介或連結院外資源9")
    outcome_col = _column_for_header(ws, "轉介或連結資源成果4")
    assert ws.cell(row=2, column=internal_col).value == "3.社福資源"
    assert ws.cell(row=2, column=external_col).value == "9.照護資訊"
    assert ws.cell(row=2, column=outcome_col).value == "4.獲得服務協助"
    wb.close()

    reopened = WorkbookWriter(working)

    expected_summary = "|".join(
        sorted(
            [
                "health_medical:screening_prevention",
                "supplies:wig_hat",
                "internal:social_welfare",
                "external:care_information",
                "outcomes:received_service_help",
            ]
        )
    )
    expected_key = ("2026-03-15", "王小明", "A123456", expected_summary)

    assert expected_key in reopened.existing_duplicate_keys()
    reopened.close()


def test_writer_raises_when_service_prefix_missing(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    wb = load_workbook(template)
    ws = wb["個案總表"]
    missing_header = "提供實體用品及設備1"
    ws.cell(row=1, column=_column_for_header(ws, missing_header)).value = None
    wb.save(template)
    wb.close()

    record = make_record()
    writer = WorkbookWriter.create_from_template(template, working)
    with pytest.raises(
        ValueError, match=re.escape("Missing workbook service columns for 提供實體用品及設備")
    ):
        writer.write_record(record)
    writer.close()


def test_writer_falls_back_to_sparse_column_when_numbered_column_absent(tmp_path: Path) -> None:
    # psychology is internal-referral option 5; its column 轉介或連結院內資源5 is absent from the
    # minimal template, so it falls back to the first free 轉介或連結院內資源* column — but the
    # written value is the real label (5.心理相關), never the raw code (#service-write-mapping).
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    record = make_record()
    record.services = Services(internal_referrals=["psychology"])

    writer = WorkbookWriter.create_from_template(template, working)
    writer.write_record(record)
    writer.save()
    writer.close()

    wb = load_workbook(working)
    ws = wb["個案總表"]
    internal_col = _column_for_header(ws, "轉介或連結院內資源3")
    assert ws.cell(row=2, column=internal_col).value == "5.心理相關"
    wb.close()

    # Duplicate detection still round-trips: the written label reverses to the same code, so the
    # re-opened workbook's summary matches Services.summary().
    reopened = WorkbookWriter(working)
    assert record.duplicate_key() in reopened.existing_duplicate_keys()
    reopened.close()


def test_writer_raises_when_sparse_service_columns_are_full(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    record = make_record()
    record.services = Services(internal_referrals=["social_welfare", "psychology"])

    writer = WorkbookWriter.create_from_template(template, working)
    with pytest.raises(ValueError, match=re.escape("No available workbook column for service psychology")):
        writer.write_record(record)
    writer.close()


def test_existing_duplicate_keys_normalize_excel_date_and_whitespace(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    record = make_record()

    writer = WorkbookWriter.create_from_template(template, working)
    writer.write_record(record)
    writer.save()
    writer.close()

    wb = load_workbook(working)
    ws = wb["個案總表"]
    service_date_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["service_date"])
    name_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["name"])
    id_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["medical_record_no"])
    ws.cell(row=2, column=service_date_col, value=datetime(2026, 3, 15))
    ws.cell(row=2, column=name_col, value=f" {record.name} ")
    ws.cell(row=2, column=id_col, value=f" {record.medical_record_no} ")
    wb.save(working)
    wb.close()

    reopened = WorkbookWriter(working)

    assert record.duplicate_key() in reopened.existing_duplicate_keys()
    reopened.close()


def test_existing_duplicate_keys_include_raw_service_codes(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    create_workbook_template(template)

    wb = load_workbook(template)
    ws = wb["個案總表"]
    ws.cell(row=2, column=_column_for_header(ws, BASIC_COLUMN_BY_FIELD["service_date"]), value="2026-04-01")
    ws.cell(row=2, column=_column_for_header(ws, BASIC_COLUMN_BY_FIELD["name"]), value="李小美")
    ws.cell(row=2, column=_column_for_header(ws, BASIC_COLUMN_BY_FIELD["medical_record_no"]), value="B987654")
    ws.cell(row=2, column=_column_for_header(ws, "諮詢-健康與醫療系統1"), value="some_code")
    ws.cell(row=2, column=_column_for_header(ws, "提供實體用品及設備1"), value="some_supply")
    ws.cell(row=2, column=_column_for_header(ws, "轉介或連結院內資源3"), value="some_resource")
    ws.cell(row=2, column=_column_for_header(ws, "轉介或連結院外資源9"), value="some_external")
    ws.cell(row=2, column=_column_for_header(ws, "轉介或連結資源成果4"), value="some_outcome")
    wb.save(template)
    wb.close()

    expected_summary = "|".join(
        sorted(
            [
                "health_medical:some_code",
                "supplies:some_supply",
                "internal:some_resource",
                "external:some_external",
                "outcomes:some_outcome",
            ]
        )
    )
    expected_key = ("2026-04-01", "李小美", "B987654", expected_summary)

    reopened = WorkbookWriter(template)

    assert expected_key in reopened.existing_duplicate_keys()
    reopened.close()


def test_existing_duplicate_keys_include_non_health_consultation_summary(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    create_workbook_template(template)

    wb = load_workbook(template)
    ws = wb["個案總表"]
    ws.cell(row=2, column=_column_for_header(ws, BASIC_COLUMN_BY_FIELD["service_date"]), value="2026-05-01")
    ws.cell(row=2, column=_column_for_header(ws, BASIC_COLUMN_BY_FIELD["name"]), value="李大明")
    ws.cell(row=2, column=_column_for_header(ws, BASIC_COLUMN_BY_FIELD["medical_record_no"]), value="C111111")
    ws.cell(row=2, column=_column_for_header(ws, "諮詢-營養與飲食1"), value="nutrition_code")
    wb.save(template)
    wb.close()

    expected_key = ("2026-05-01", "李大明", "C111111", "nutrition_diet:nutrition_code")

    reopened = WorkbookWriter(template)

    assert expected_key in reopened.existing_duplicate_keys()
    reopened.close()


def test_workbook_writer_close_closes_workbook(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    create_workbook_template(template)
    writer = WorkbookWriter(template)
    real_close = writer.workbook.close
    called = False

    def fake_close() -> None:
        nonlocal called
        called = True
        real_close()

    writer.workbook.close = fake_close  # type: ignore[method-assign]

    writer.close()

    assert called is True


def test_workbook_writer_context_manager_closes_workbook(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    create_workbook_template(template)
    writer = WorkbookWriter(template)
    real_close = writer.workbook.close
    called = False

    def fake_close() -> None:
        nonlocal called
        called = True
        real_close()

    writer.workbook.close = fake_close  # type: ignore[method-assign]

    with writer as managed:
        assert managed is writer

    assert called is True


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        (True, "是"),
        (False, "否"),
        (None, None),
    ],
)
def test_writer_sets_discharge_followup(
    tmp_path: Path, value: bool | None, expected: str | None
) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    record = make_record()
    record.discharge_followup = value

    writer = WorkbookWriter.create_from_template(template, working)
    writer.write_record(record)
    writer.save()
    writer.close()

    wb = load_workbook(working)
    ws = wb["個案總表"]
    column = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["discharge_followup"])

    assert ws.cell(row=2, column=column).value == expected
    wb.close()


def test_missing_patient_header_raises_value_error(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    create_workbook_template(template)

    wb = load_workbook(template)
    ws = wb["個案總表"]
    missing_header = BASIC_COLUMN_BY_FIELD["nationality"]
    missing_cell = ws.cell(row=1, column=_column_for_header(ws, missing_header))
    missing_cell.value = None
    wb.save(template)
    wb.close()

    with pytest.raises(ValueError, match=re.escape(missing_header)):
        WorkbookWriter(template)


def test_writer_leaves_newly_diagnosed_blank_when_none(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    record = make_record()
    record.patient_fields.newly_diagnosed_within_year = None

    writer = WorkbookWriter.create_from_template(template, working)
    writer.write_record(record)
    writer.save()
    writer.close()

    wb = load_workbook(working)
    ws = wb["個案總表"]
    column = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["newly_diagnosed_within_year"])

    assert ws.cell(row=2, column=column).value is None
    wb.close()


def test_write_record_to_explicit_row_overwrites_without_appending(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    writer = WorkbookWriter.create_from_template(template, working)
    try:
        first = make_record("r1")
        first.name = "王小明"
        first.medical_record_no = "A1"
        second = make_record("r2")
        second.name = "李大華"
        second.medical_record_no = "B2"
        row1 = writer.write_record(first)
        row2 = writer.write_record(second)
        assert row2 == row1 + 1
        corrected = make_record("r1b")
        corrected.name = "王小華"
        corrected.medical_record_no = "A9"
        out = writer.write_record(corrected, row=row1)
        assert out == row1
        writer.save()
    finally:
        writer.close()

    wb = load_workbook(working)
    try:
        sheet = wb["個案總表"]
        name_col = _column_for_header(sheet, BASIC_COLUMN_BY_FIELD["name"])
        mrn_col = _column_for_header(sheet, BASIC_COLUMN_BY_FIELD["medical_record_no"])
        assert sheet.cell(row=row1, column=name_col).value == "王小華"
        assert sheet.cell(row=row1, column=mrn_col).value == "A9"
        assert sheet.cell(row=row2, column=name_col).value == "李大華"
    finally:
        wb.close()


def test_overwrite_clears_stale_service_cells(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    writer = WorkbookWriter.create_from_template(template, working)
    try:
        full = make_record("r1")
        full.name = "王小明"
        full.medical_record_no = "A1"
        full.services.consultation["health_medical"] = ["screening_prevention"]
        row = writer.write_record(full)
        empty = make_record("r1b")
        empty.name = "王小明"
        empty.medical_record_no = "A1"
        empty.services = Services()
        writer.write_record(empty, row=row)
        writer.save()
    finally:
        writer.close()

    wb = load_workbook(working)
    try:
        sheet = wb["個案總表"]
        values = [cell.value for cell in sheet[row] if cell.value not in (None, "")]
        assert "1.癌症篩檢與預防" not in values
    finally:
        wb.close()


def test_every_service_option_writes_correct_label_to_its_numbered_column(tmp_path: Path) -> None:
    """#service-write-mapping: every selected service option must land in its OWN numbered
    column (prefix+number) with the real form label — never the raw English code in the first
    free column (the prior bug, e.g. fatigue_strength → 諮詢-症狀與副作用照護1 = 'fatigue_strength')."""
    from ocr_from2xlsx.domain import Services
    from ocr_from2xlsx.form_layout import service_record_layout
    from tests.fixtures import create_full_workbook_template

    layout = service_record_layout()
    group_prefix = {  # record_path -> 個案總表 column prefix
        "services.consultation.health_medical": "諮詢-健康與醫療系統",
        "services.consultation.symptom_side_effect": "諮詢-症狀與副作用照護",
        "services.consultation.nutrition_diet": "諮詢-營養與飲食",
        "services.consultation.psychosocial_emotion": "諮詢-社會心理情緒",
        "services.consultation.financial_social": "諮詢-經濟與社會資源",
        "services.consultation.care_support": "諮詢-照顧與支持",
        "services.supplies": "提供實體用品及設備",
        "services.internal_referrals": "轉介或連結院內資源",
        "services.external_referrals": "轉介或連結院外資源",
        "services.referral_outcomes": "轉介或連結資源成果",
    }
    fields = {p: layout.field_by_key(p[len("services."):]) for p in group_prefix}

    # Select EVERY option in every service field.
    record = make_record()
    record.services = Services(
        consultation={
            p.rsplit(".", 1)[-1]: [o.code for o in f.options]
            for p, f in fields.items()
            if p.startswith("services.consultation.")
        },
        supplies=[o.code for o in fields["services.supplies"].options],
        internal_referrals=[o.code for o in fields["services.internal_referrals"].options],
        external_referrals=[o.code for o in fields["services.external_referrals"].options],
        referral_outcomes=[o.code for o in fields["services.referral_outcomes"].options],
    )

    template = tmp_path / "t.xlsx"
    working = tmp_path / "w.xlsx"
    create_full_workbook_template(template)
    writer = WorkbookWriter.create_from_template(template, working)
    writer.write_record(record)
    writer.save()
    writer.close()

    wb = load_workbook(working)
    try:
        sheet = wb["個案總表"]
        all_codes: set[str] = set()
        for path, prefix in group_prefix.items():
            for i, opt in enumerate(fields[path].options, start=1):
                all_codes.add(opt.code)
                col = _column_for_header(sheet, f"{prefix}{i}")
                assert sheet.cell(row=2, column=col).value == opt.label, (
                    f"{path}#{i} ({opt.code}) -> {prefix}{i} should be {opt.label!r}, "
                    f"got {sheet.cell(row=2, column=col).value!r}"
                )
        # The exact case from the bug report: fatigue_strength -> AE / 諮詢-症狀與副作用照護4.
        ae = _column_for_header(sheet, "諮詢-症狀與副作用照護4")
        assert sheet.cell(row=2, column=ae).value == "4.疲憊與體力"
        # No raw English code leaked into any cell.
        for col in range(1, sheet.max_column + 1):
            val = sheet.cell(row=2, column=col).value
            assert val not in all_codes, f"raw code {val!r} leaked into column {col}"
    finally:
        wb.close()
