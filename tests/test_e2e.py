from __future__ import annotations

import json
from pathlib import Path
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

import ocr_from2xlsx.prepare_records as prepare_records_module
from ocr_from2xlsx.constants import (
    BASIC_COLUMN_BY_FIELD,
    GENDER_LABELS,
    IDENTITY_LABELS,
    WORKBOOK_SHEET,
)
from ocr_from2xlsx.cli import main
from ocr_from2xlsx.json_io import load_batch
from ocr_from2xlsx.session import ImportSession
from tests.fixtures import create_workbook_template


def _column_for_header(sheet, header: str) -> int:
    for cell in sheet[1]:
        if cell.value == header:
            return cell.column
    raise AssertionError(f"Missing header in fixture: {header}")


def _non_empty_rows(sheet, column: int) -> list[int]:
    rows: list[int] = []
    for row in range(2, sheet.max_row + 1):
        if sheet.cell(row=row, column=column).value not in (None, ""):
            rows.append(row)
    return rows


def _load_workbook(path: Path):
    keep_vba = path.suffix.lower() in {".xlsm", ".xltm"}
    return load_workbook(path, keep_vba=keep_vba)


class _FixedDatetime:
    def __init__(self, fixed_now: datetime) -> None:
        self._fixed_now = fixed_now

    def now(self):
        return self

    def astimezone(self):
        return self._fixed_now


def test_end_to_end_import(tmp_path: Path) -> None:
    repo_root = Path(__file__).resolve().parents[1]
    sample_path = repo_root / "src" / "ocr_from2xlsx" / "sample_data.json"
    batch = load_batch(sample_path)

    template_path = tmp_path / "template.xlsx"
    working_path = tmp_path / "working.xlsx"
    create_workbook_template(template_path)

    template_wb = _load_workbook(template_path)
    working_wb = None
    try:
        template_ws = template_wb[WORKBOOK_SHEET]

        with ImportSession.start(template_path, working_path) as session:
            results = session.accept_scan_batch(batch)

        working_wb = _load_workbook(working_path)
        working_ws = working_wb[WORKBOOK_SHEET]

        assert working_wb.sheetnames == template_wb.sheetnames

        assert working_ws["A1"].fill.fill_type == template_ws["A1"].fill.fill_type
        assert working_ws["A1"].fill.fgColor.value == template_ws["A1"].fill.fgColor.value
        assert working_ws.column_dimensions["A"].width == template_ws.column_dimensions["A"].width
        assert working_ws.column_dimensions["B"].width == template_ws.column_dimensions["B"].width
        assert working_wb["一月"]["A1"].value == template_wb["一月"]["A1"].value

        name_col = _column_for_header(working_ws, BASIC_COLUMN_BY_FIELD["name"])
        written_results = [result for result in results if result.status in {"forced", "written"}]
        assert written_results

        rows_with_names = _non_empty_rows(working_ws, name_col)
        assert len(rows_with_names) == len(written_results)

        record_by_id = {record.record_id: record for record in batch.records}
        first_written = written_results[0]
        row = first_written.row_number
        assert row is not None
        record = record_by_id[first_written.record_id]

        assert working_ws.cell(row=row, column=name_col).value == record.name
        id_col = _column_for_header(working_ws, BASIC_COLUMN_BY_FIELD["medical_record_no"])
        assert working_ws.cell(row=row, column=id_col).value == record.medical_record_no
        date_col = _column_for_header(working_ws, BASIC_COLUMN_BY_FIELD["service_date"])
        assert working_ws.cell(row=row, column=date_col).value == record.service_date
        identity_col = _column_for_header(working_ws, BASIC_COLUMN_BY_FIELD["identity"])
        assert working_ws.cell(row=row, column=identity_col).value == IDENTITY_LABELS[record.identity]
        gender_col = _column_for_header(working_ws, BASIC_COLUMN_BY_FIELD["gender"])
        assert working_ws.cell(row=row, column=gender_col).value == GENDER_LABELS[record.gender]

        template_vba = getattr(template_wb, "vba_archive", None)
        if template_vba is not None:
            working_vba = getattr(working_wb, "vba_archive", None)
            assert working_vba is not None
            assert working_vba.namelist() == template_vba.namelist()
    finally:
        template_wb.close()
        if working_wb is not None:
            working_wb.close()


def test_end_to_end_prepare_records_then_import_json(tmp_path: Path) -> None:
    fixture_dir = Path(__file__).parent / "fixtures" / "pdf"
    expected_path = fixture_dir / "for testing only.expected.json"
    prepared_json = tmp_path / "prepared.json"
    template_path = tmp_path / "template.xlsx"
    working_path = tmp_path / "working.xlsx"
    report_json = tmp_path / "report.json"
    report_csv = tmp_path / "report.csv"
    create_workbook_template(template_path)
    fixed_now = datetime(2026, 5, 26, 0, 0, 0, tzinfo=timezone(timedelta(hours=8)))
    original_datetime = prepare_records_module.datetime

    prepare_records_module.datetime = _FixedDatetime(fixed_now)
    try:
        assert (
            main(
                [
                    "prepare-records",
                    "--input",
                    str(fixture_dir / "for testing only.pdf"),
                    "--output",
                    str(prepared_json),
                    "--ocr-fixture",
                    str(fixture_dir / "for testing only.ocr.json"),
                ]
            )
            == 0
        )
    finally:
        prepare_records_module.datetime = original_datetime

    assert json.loads(prepared_json.read_text(encoding="utf-8")) == json.loads(
        expected_path.read_text(encoding="utf-8")
    )

    assert (
        main(
            [
                "import-json",
                "--input",
                str(prepared_json),
                "--template",
                str(template_path),
                "--working",
                str(working_path),
                "--report-json",
                str(report_json),
                "--report-csv",
                str(report_csv),
            ]
        )
        == 0
    )

    wb = _load_workbook(working_path)
    try:
        ws = wb[WORKBOOK_SHEET]
        name_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["name"])
        assert ws.cell(row=2, column=name_col).value == "AI test"
    finally:
        wb.close()
