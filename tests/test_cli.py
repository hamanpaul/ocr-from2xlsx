from __future__ import annotations

import json
from pathlib import Path

import fitz
import pytest
from openpyxl import load_workbook

from ocr_from2xlsx.cli import _resolve_name_crop_path, build_parser, main
from ocr_from2xlsx.constants import BASIC_COLUMN_BY_FIELD, WORKBOOK_SHEET
from ocr_from2xlsx.domain import Batch, SourceBatch
from ocr_from2xlsx.json_io import dump_batch
from ocr_from2xlsx.session import ImportSession
from tests.fixtures import create_workbook_template
from tests.test_json_io import make_record


def test_import_json_cli_writes_workbook(tmp_path: Path, capsys) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    input_json = tmp_path / "records.json"
    report_json = tmp_path / "report.json"
    report_csv = tmp_path / "report.csv"
    create_workbook_template(template)
    dump_batch(
        Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="json_import",
                template_name="template.xlsx",
            ),
            records=[make_record()],
        ),
        input_json,
    )

    exit_code = main(
        [
            "import-json",
            "--input",
            str(input_json),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(report_json),
            "--report-csv",
            str(report_csv),
        ]
    )

    assert exit_code == 0
    captured = capsys.readouterr()
    assert captured.out == f"{working}\n"
    assert working.exists()
    wb = load_workbook(working)
    ws = wb["個案總表"]
    assert ws["E2"].value == "王小明"
    wb.close()
    assert report_json.exists()
    assert report_csv.exists()
    report = json.loads(report_json.read_text(encoding="utf-8"))
    assert len(report) == 1
    assert report[0]["status"] == "written"


def test_import_json_cli_reports_blocked_records(tmp_path: Path, capsys) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    input_json = tmp_path / "records.json"
    report_json = tmp_path / "report.json"
    report_csv = tmp_path / "report.csv"
    create_workbook_template(template)
    blocked_record = make_record()
    blocked_record.service_date = ""
    dump_batch(
        Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="json_import",
                template_name="template.xlsx",
            ),
            records=[blocked_record],
        ),
        input_json,
    )

    exit_code = main(
        [
            "import-json",
            "--input",
            str(input_json),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(report_json),
            "--report-csv",
            str(report_csv),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 1
    assert captured.out == f"{working}\n"
    assert working.exists()
    assert report_json.exists()
    assert report_csv.exists()
    report = json.loads(report_json.read_text(encoding="utf-8"))
    assert report[0]["status"] == "blocked"


def test_import_json_cli_reports_input_error_without_traceback(
    tmp_path: Path, capsys
) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    missing_input = tmp_path / "missing.json"
    report_json = tmp_path / "report.json"
    report_csv = tmp_path / "report.csv"
    create_workbook_template(template)

    exit_code = main(
        [
            "import-json",
            "--input",
            str(missing_input),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(report_json),
            "--report-csv",
            str(report_csv),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")
    assert "Traceback" not in captured.err


def test_import_json_cli_warns_when_report_write_fails(
    tmp_path: Path, capsys
) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    input_json = tmp_path / "records.json"
    report_json = tmp_path / "report-as-directory"
    report_csv = tmp_path / "report.csv"
    create_workbook_template(template)
    report_json.mkdir()
    dump_batch(
        Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="json_import",
                template_name="template.xlsx",
            ),
            records=[make_record()],
        ),
        input_json,
    )

    exit_code = main(
        [
            "import-json",
            "--input",
            str(input_json),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(report_json),
            "--report-csv",
            str(report_csv),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert working.exists()
    assert "working XLSX may contain imported records" in captured.err
    assert "Traceback" not in captured.err


def test_import_json_cli_omits_import_warning_when_empty_report_write_fails(
    tmp_path: Path, capsys
) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    input_json = tmp_path / "records.json"
    report_json = tmp_path / "report-as-directory"
    report_csv = tmp_path / "report.csv"
    create_workbook_template(template)
    report_json.mkdir()
    dump_batch(
        Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="json_import",
                template_name="template.xlsx",
            ),
            records=[],
        ),
        input_json,
    )

    exit_code = main(
        [
            "import-json",
            "--input",
            str(input_json),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(report_json),
            "--report-csv",
            str(report_csv),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "report writing did not complete" in captured.err
    assert "working XLSX may contain imported records" not in captured.err
    assert "Traceback" not in captured.err


def test_import_json_cli_omits_import_warning_when_blocked_report_write_fails(
    tmp_path: Path, capsys
) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    input_json = tmp_path / "records.json"
    report_json = tmp_path / "report-as-directory"
    report_csv = tmp_path / "report.csv"
    create_workbook_template(template)
    report_json.mkdir()
    blocked_record = make_record()
    blocked_record.service_date = ""
    dump_batch(
        Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="json_import",
                template_name="template.xlsx",
            ),
            records=[blocked_record],
        ),
        input_json,
    )

    exit_code = main(
        [
            "import-json",
            "--input",
            str(input_json),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(report_json),
            "--report-csv",
            str(report_csv),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "report writing did not complete" in captured.err
    assert "working XLSX may contain imported records" not in captured.err
    assert "Traceback" not in captured.err


def test_import_json_allow_incomplete_writes_forced_row(tmp_path: Path) -> None:
    import json as _json

    template = tmp_path / "t.xlsx"
    working = tmp_path / "w.xlsx"
    create_workbook_template(template)

    batch = {
        "schema_version": "service_record.v1",
        "source_batch": {
            "created_at": "2026-05-26T00:00:00+08:00",
            "source_type": "manual",
            "template_name": "t",
        },
        "records": [
            {
                "record_id": "pdf-0001",
                "service_date": "2025-06-25",
                "identity": "patient",
                "name": "葉心安",
                "medical_record_no": "6250712919",
                "gender": "female",
            }
        ],
    }
    inp = tmp_path / "in.json"
    inp.write_text(_json.dumps(batch), encoding="utf-8")

    code = main(
        [
            "import-json",
            "--input",
            str(inp),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(tmp_path / "r.json"),
            "--report-csv",
            str(tmp_path / "r.csv"),
            "--allow-incomplete",
        ]
    )

    assert code == 0
    wb = load_workbook(working)
    ws = wb[WORKBOOK_SHEET]
    name_col = next(c.column for c in ws[1] if c.value == BASIC_COLUMN_BY_FIELD["name"])
    assert ws.cell(row=2, column=name_col).value == "葉心安"
    wb.close()


def _make_unconfirmed_name_batch() -> dict:
    """Build a minimal batch JSON with a public_other record carrying name.unconfirmed."""
    return {
        "schema_version": "service_record.v1",
        "source_batch": {
            "created_at": "2026-05-31T00:00:00+08:00",
            "source_type": "manual",
            "template_name": "t",
        },
        "records": [
            {
                "record_id": "pdf-0001",
                "service_date": "2025-06-25",
                "identity": "public_other",
                "name": "葉心安",
                "medical_record_no": "6250712919",
                "gender": "female",
                "review": {"status": "pending"},
                "ocr": {"raw_text": "", "warnings": ["name.unconfirmed"]},
            }
        ],
    }


def test_import_json_allow_unconfirmed_name_blocked_without_flag(tmp_path: Path) -> None:
    import json as _json

    template = tmp_path / "t.xlsx"
    working = tmp_path / "w.xlsx"
    create_workbook_template(template)

    inp = tmp_path / "in.json"
    inp.write_text(_json.dumps(_make_unconfirmed_name_batch()), encoding="utf-8")

    code = main(
        [
            "import-json",
            "--input",
            str(inp),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(tmp_path / "r.json"),
            "--report-csv",
            str(tmp_path / "r.csv"),
        ]
    )

    assert code == 1  # blocked
    wb = load_workbook(working)
    ws = wb[WORKBOOK_SHEET]
    name_col = next(c.column for c in ws[1] if c.value == BASIC_COLUMN_BY_FIELD["name"])
    assert ws.cell(row=2, column=name_col).value is None  # NOT written
    wb.close()


def test_import_json_allow_unconfirmed_name_writes_and_keeps_warning(tmp_path: Path) -> None:
    import json as _json

    template = tmp_path / "t.xlsx"
    working = tmp_path / "w.xlsx"
    create_workbook_template(template)

    inp = tmp_path / "in.json"
    inp.write_text(_json.dumps(_make_unconfirmed_name_batch()), encoding="utf-8")
    report_json = tmp_path / "r.json"

    code = main(
        [
            "import-json",
            "--input",
            str(inp),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(report_json),
            "--report-csv",
            str(tmp_path / "r.csv"),
            "--allow-unconfirmed-name",
        ]
    )

    assert code == 0
    wb = load_workbook(working)
    ws = wb[WORKBOOK_SHEET]
    name_col = next(c.column for c in ws[1] if c.value == BASIC_COLUMN_BY_FIELD["name"])
    assert ws.cell(row=2, column=name_col).value == "葉心安"  # written
    wb.close()
    report = _json.loads(report_json.read_text(encoding="utf-8"))
    assert report[0]["status"] in ("written", "forced")
    assert "name.unconfirmed" in report[0]["warnings"]  # warning retained


def test_import_json_cli_warns_when_accept_fails_before_returning_result(
    tmp_path: Path, capsys, monkeypatch: pytest.MonkeyPatch
) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    input_json = tmp_path / "records.json"
    report_json = tmp_path / "report.json"
    report_csv = tmp_path / "report.csv"
    create_workbook_template(template)
    dump_batch(
        Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="json_import",
                template_name="template.xlsx",
            ),
            records=[make_record()],
        ),
        input_json,
    )

    class SaveFailingSession:
        closed = False

        def __enter__(self) -> "SaveFailingSession":
            return self

        def __exit__(self, exc_type: object, exc: object, traceback: object) -> None:
            self.closed = True

        def accept_scan(self, record: object, force: bool = False, allow_unconfirmed_name: bool = False) -> None:
            raise OSError("save failed")

        def write_report(self, json_path: object, csv_path: object) -> None:
            raise AssertionError("report should not be written")

    failing_session = SaveFailingSession()
    monkeypatch.setattr(
        ImportSession,
        "start",
        staticmethod(lambda template_path, working_path: failing_session),
    )

    exit_code = main(
        [
            "import-json",
            "--input",
            str(input_json),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(report_json),
            "--report-csv",
            str(report_csv),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "save failed" in captured.err
    assert "working XLSX may contain imported records" in captured.err
    assert "Traceback" not in captured.err
    assert failing_session.closed


def test_import_json_cli_warns_when_accept_fails_after_write(
    tmp_path: Path, capsys
) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    input_json = tmp_path / "records.json"
    report_json = tmp_path / "report.json"
    report_csv = tmp_path / "report.csv"
    create_workbook_template(template)
    first_record = make_record("scan-0001")
    failing_record = make_record("scan-0002")
    failing_record.name = "王大明"
    failing_record.medical_record_no = "B234567"
    failing_record.services.consultation["health_medical"] = [
        "screening_prevention",
        "disease_treatment_knowledge",
    ]
    dump_batch(
        Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="json_import",
                template_name="template.xlsx",
            ),
            records=[first_record, failing_record],
        ),
        input_json,
    )

    exit_code = main(
        [
            "import-json",
            "--input",
            str(input_json),
            "--template",
            str(template),
            "--working",
            str(working),
            "--report-json",
            str(report_json),
            "--report-csv",
            str(report_csv),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert working.exists()
    assert "working XLSX may contain imported records" in captured.err
    assert "Traceback" not in captured.err


def test_prepare_records_cli_writes_batch_json_from_pdf_fixture(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch, capsys
) -> None:
    fixture_dir = Path(__file__).parent / "fixtures" / "pdf"
    output_json = tmp_path / "prepared.json"

    exit_code = main(
        [
            "prepare-records",
            "--input",
            str(fixture_dir / "for testing only.pdf"),
            "--output",
            str(output_json),
            "--ocr-fixture",
            str(fixture_dir / "for testing only.ocr.json"),
        ]
    )

    assert exit_code == 0
    assert output_json.exists()
    captured = capsys.readouterr()
    assert captured.out == f"{output_json}\n"


def test_prepare_records_help_mentions_pdf_inputs_only(capsys) -> None:
    parser = build_parser()

    with pytest.raises(SystemExit):
        parser.parse_args(["prepare-records", "--help"])

    captured = capsys.readouterr()
    help_text = captured.out
    assert "Input PDF path." in help_text
    assert "Input PDF or image path." not in help_text


def test_root_help_mentions_pdf_prep_workflow(capsys) -> None:
    parser = build_parser()

    assert parser.description is not None
    assert "PDF" in parser.description

    with pytest.raises(SystemExit):
        parser.parse_args(["--help"])

    capsys.readouterr()


def test_prepare_records_cli_requires_ocr_fixture(
    tmp_path: Path, capsys
) -> None:
    fixture_dir = Path(__file__).parent / "fixtures" / "pdf"
    output_json = tmp_path / "prepared.json"

    exit_code = main(
        [
            "prepare-records",
            "--input",
            str(fixture_dir / "for testing only.pdf"),
            "--output",
            str(output_json),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "--ocr-fixture" in captured.err


def _install_echo_plugin(tmp_path):
    import json as _json
    import shutil as _shutil
    import sys as _sys
    from pathlib import Path as _Path

    fixture = _Path(__file__).parent / "fixtures" / "plugin"
    plugin_dir = tmp_path / "plugin"
    plugin_dir.mkdir()
    _shutil.copy(fixture / "echo_plugin.py", plugin_dir / "echo_plugin.py")
    (plugin_dir / "plugin.json").write_text(
        _json.dumps(
            {"contract_version": "ocr_plugin.v1", "command": [_sys.executable, "echo_plugin.py"]}
        ),
        encoding="utf-8",
    )
    return plugin_dir


def test_prepare_records_with_plugin_backend(tmp_path):
    import json as _json
    from pathlib import Path as _Path

    from ocr_from2xlsx.cli import main

    plugin_dir = _install_echo_plugin(tmp_path)
    pdf = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.pdf"
    output = tmp_path / "prepared.json"

    code = main(
        [
            "prepare-records",
            "--input",
            str(pdf),
            "--output",
            str(output),
            "--ocr-backend",
            "plugin",
            "--ocr-plugin-dir",
            str(plugin_dir),
        ]
    )

    assert code == 0
    data = _json.loads(output.read_text(encoding="utf-8"))
    assert data["records"][0]["name"] == "Plugin Echo"


def test_prepare_records_plugin_missing_reports_error(tmp_path, capsys):
    from pathlib import Path as _Path

    from ocr_from2xlsx.cli import main

    pdf = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.pdf"
    output = tmp_path / "prepared.json"

    code = main(
        [
            "prepare-records",
            "--input",
            str(pdf),
            "--output",
            str(output),
            "--ocr-backend",
            "plugin",
            "--ocr-plugin-dir",
            str(tmp_path / "no-plugin-here"),
        ]
    )

    assert code == 2
    assert "plugin" in capsys.readouterr().err.lower()


def test_prepare_records_fixture_backend_still_default(tmp_path):
    import json as _json
    from pathlib import Path as _Path

    from ocr_from2xlsx.cli import main

    pdf = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.pdf"
    fixture = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.ocr.json"
    output = tmp_path / "prepared.json"

    code = main(
        [
            "prepare-records",
            "--input",
            str(pdf),
            "--output",
            str(output),
            "--ocr-fixture",
            str(fixture),
        ]
    )

    assert code == 0
    data = _json.loads(output.read_text(encoding="utf-8"))
    assert data["records"][0]["name"] == "AI test"


def test_prepare_records_cli_rejects_unknown_template_id(
    tmp_path: Path, capsys
) -> None:
    fixture_dir = Path(__file__).parent / "fixtures" / "pdf"
    output_json = tmp_path / "prepared.json"

    exit_code = main(
        [
            "prepare-records",
            "--input",
            str(fixture_dir / "for testing only.pdf"),
            "--output",
            str(output_json),
            "--ocr-fixture",
            str(fixture_dir / "for testing only.ocr.json"),
            "--template-id",
            "service_record.v2",
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert "Unsupported template_id: 'service_record.v2'" in captured.err
    assert "Traceback" not in captured.err


def test_prepare_records_cli_reports_missing_input_without_traceback(
    tmp_path: Path, capsys
) -> None:
    fixture_dir = Path(__file__).parent / "fixtures" / "pdf"
    output_json = tmp_path / "prepared.json"
    missing_input = tmp_path / "missing.pdf"

    exit_code = main(
        [
            "prepare-records",
            "--input",
            str(missing_input),
            "--output",
            str(output_json),
            "--ocr-fixture",
            str(fixture_dir / "for testing only.ocr.json"),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")
    assert "Traceback" not in captured.err


def test_prepare_records_cli_reports_missing_ocr_fixture_without_traceback(
    tmp_path: Path, capsys
) -> None:
    fixture_dir = Path(__file__).parent / "fixtures" / "pdf"
    output_json = tmp_path / "prepared.json"
    missing_fixture = tmp_path / "missing.ocr.json"

    exit_code = main(
        [
            "prepare-records",
            "--input",
            str(fixture_dir / "for testing only.pdf"),
            "--output",
            str(output_json),
            "--ocr-fixture",
            str(missing_fixture),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")
    assert "Traceback" not in captured.err


def test_prepare_records_cli_reports_malformed_ocr_fixture_without_traceback(
    tmp_path: Path, capsys
) -> None:
    fixture_dir = Path(__file__).parent / "fixtures" / "pdf"
    output_json = tmp_path / "prepared.json"
    malformed_fixture = tmp_path / "malformed.ocr.json"
    malformed_fixture.write_text("{\"unexpected\": []}", encoding="utf-8")

    exit_code = main(
        [
            "prepare-records",
            "--input",
            str(fixture_dir / "for testing only.pdf"),
            "--output",
            str(output_json),
            "--ocr-fixture",
            str(malformed_fixture),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")
    assert "Traceback" not in captured.err


def test_prepare_records_cli_reports_non_object_source_without_traceback(
    tmp_path: Path, capsys
) -> None:
    fixture_dir = Path(__file__).parent / "fixtures" / "pdf"
    output_json = tmp_path / "prepared.json"
    malformed_fixture = tmp_path / "malformed-source.ocr.json"
    malformed_fixture.write_text(
        json.dumps(
            {
                "pages": [
                    {
                        "document_name": "for testing only.pdf",
                        "page_number": 1,
                        "record": {
                            "record_id": "pdf-0001",
                            "service_date": "2026-05-26",
                            "identity": "patient",
                            "name": "AI test",
                            "medical_record_no": "TRAINING-ONLY",
                            "gender": "female",
                            "source": "broken",
                        },
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    exit_code = main(
        [
            "prepare-records",
            "--input",
            str(fixture_dir / "for testing only.pdf"),
            "--output",
            str(output_json),
            "--ocr-fixture",
            str(malformed_fixture),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")
    assert "Traceback" not in captured.err


def test_prepare_records_cli_reports_null_source_without_traceback(
    tmp_path: Path, capsys
) -> None:
    fixture_dir = Path(__file__).parent / "fixtures" / "pdf"
    output_json = tmp_path / "prepared.json"
    malformed_fixture = tmp_path / "null-source.ocr.json"
    malformed_fixture.write_text(
        json.dumps(
            {
                "pages": [
                    {
                        "document_name": "for testing only.pdf",
                        "page_number": 1,
                        "record": {
                            "record_id": "pdf-0001",
                            "service_date": "2026-05-26",
                            "identity": "patient",
                            "name": "AI test",
                            "medical_record_no": "TRAINING-ONLY",
                            "gender": "female",
                            "source": None,
                        },
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    exit_code = main(
        [
            "prepare-records",
            "--input",
            str(fixture_dir / "for testing only.pdf"),
            "--output",
            str(output_json),
            "--ocr-fixture",
            str(malformed_fixture),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")
    assert "Traceback" not in captured.err


def test_prepare_records_name_agent_absent_is_noop(tmp_path):
    import json as _json
    from pathlib import Path as _Path
    from ocr_from2xlsx.cli import main

    def _prepare_records_from_paths(*args, **kwargs):
        record = make_record()
        record.name = ""
        record.ocr.warnings = []
        return Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="prepare_records",
                template_name="service_record.v1",
            ),
            records=[record],
        )

    pdf = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.pdf"
    fixture = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.ocr.json"
    out = tmp_path / "prepared.json"

    with pytest.MonkeyPatch.context() as monkeypatch:
        monkeypatch.setattr(
            "ocr_from2xlsx.prepare_records.prepare_records_from_paths",
            _prepare_records_from_paths,
        )
        code = main(
            [
                "prepare-records",
                "--input",
                str(pdf),
                "--output",
                str(out),
                "--ocr-fixture",
                str(fixture),
            ]
        )

    assert code == 0
    data = _json.loads(out.read_text(encoding="utf-8"))
    assert data["records"][0]["name"] == ""
    assert data["records"][0]["ocr"]["warnings"] == []


def test_prepare_records_disabled_name_agent_config_is_noop(tmp_path):
    import json as _json
    from pathlib import Path as _Path
    from ocr_from2xlsx.cli import main

    def _prepare_records_from_paths(*args, **kwargs):
        record = make_record()
        record.name = ""
        record.ocr.warnings = []
        return Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="prepare_records",
                template_name="service_record.v1",
            ),
            records=[record],
        )

    cfg = tmp_path / "name_agent.toml"
    cfg.write_text("enabled = false\n", encoding="utf-8")
    pdf = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.pdf"
    fixture = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.ocr.json"
    out = tmp_path / "prepared.json"

    with pytest.MonkeyPatch.context() as monkeypatch:
        monkeypatch.setattr(
            "ocr_from2xlsx.prepare_records.prepare_records_from_paths",
            _prepare_records_from_paths,
        )
        code = main(
            [
                "prepare-records",
                "--input",
                str(pdf),
                "--output",
                str(out),
                "--ocr-fixture",
                str(fixture),
                "--name-agent-config",
                str(cfg),
            ]
        )
    assert code == 0
    data = _json.loads(out.read_text(encoding="utf-8"))
    assert data["records"][0]["name"] == ""
    assert data["records"][0]["ocr"]["warnings"] == []


def test_prepare_records_unsupported_name_agent_config_is_noop(tmp_path):
    import json as _json
    from pathlib import Path as _Path

    from ocr_from2xlsx.cli import main

    record = make_record()
    record.name = ""
    record.ocr.raw_text = "王小明"
    record.ocr.warnings = []
    record.source.preprocessed_image_path = "for testing only-page-0001.png"

    def _prepare_records_from_paths(*args, **kwargs):
        return Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="prepare_records",
                template_name="service_record.v1",
            ),
            records=[record],
        )

    cfg = tmp_path / "name_agent.toml"
    cfg.write_text('enabled = true\nprovider = "nope"\n', encoding="utf-8")
    crop_path = tmp_path / "for testing only-page-0001-name.png"
    crop_path.write_bytes(b"fake png bytes")
    pdf = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.pdf"
    fixture = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.ocr.json"
    out = tmp_path / "prepared.json"

    with pytest.MonkeyPatch.context() as monkeypatch:
        monkeypatch.setattr(
            "ocr_from2xlsx.prepare_records.prepare_records_from_paths",
            _prepare_records_from_paths,
        )
        code = main(
            [
                "prepare-records",
                "--input",
                str(pdf),
                "--output",
                str(out),
                "--ocr-fixture",
                str(fixture),
                "--name-agent-config",
                str(cfg),
            ]
        )

    assert code == 0
    data = _json.loads(out.read_text(encoding="utf-8"))
    assert data["records"][0]["name"] == ""
    assert data["records"][0]["ocr"]["warnings"] == []


def test_prepare_records_effectively_null_name_agent_config_is_noop(tmp_path):
    import json as _json
    from pathlib import Path as _Path

    from ocr_from2xlsx.cli import main

    record = make_record()
    record.name = ""
    record.ocr.raw_text = "王小明"
    record.ocr.warnings = []
    record.source.preprocessed_image_path = "for testing only-page-0001.png"

    def _prepare_records_from_paths(*args, **kwargs):
        return Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="prepare_records",
                template_name="service_record.v1",
            ),
            records=[record],
        )

    cfg = tmp_path / "name_agent.toml"
    cfg.write_text('enabled = true\nprovider = "claude"\nmodel = "claude-x"\n', encoding="utf-8")
    crop_path = tmp_path / "for testing only-page-0001-name.png"
    crop_path.write_bytes(b"fake png bytes")
    pdf = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.pdf"
    fixture = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.ocr.json"
    out = tmp_path / "prepared.json"

    with pytest.MonkeyPatch.context() as monkeypatch:
        monkeypatch.setattr(
            "ocr_from2xlsx.prepare_records.prepare_records_from_paths",
            _prepare_records_from_paths,
        )
        code = main(
            [
                "prepare-records",
                "--input",
                str(pdf),
                "--output",
                str(out),
                "--ocr-fixture",
                str(fixture),
                "--name-agent-config",
                str(cfg),
            ]
        )

    assert code == 0
    data = _json.loads(out.read_text(encoding="utf-8"))
    assert data["records"][0]["name"] == ""
    assert data["records"][0]["ocr"]["warnings"] == []


def test_prepare_records_enabled_name_agent_suggests_name_from_preprocessed_crop(tmp_path):
    import json as _json
    from pathlib import Path as _Path
    from ocr_from2xlsx.correction_store import Correction, append_correction

    from ocr_from2xlsx.cli import main

    record = make_record()
    record.name = ""
    record.ocr.warnings = ["existing-warning"]
    record.source.preprocessed_image_path = "for testing only-page-0001.png"

    def _prepare_records_from_paths(*args, **kwargs):
        return Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="prepare_records",
                template_name="service_record.v1",
            ),
            records=[record],
        )

    class FakeAgent:
        pass

    fake_agent = FakeAgent()
    observed: dict[str, object] = {}

    def _build_agent(config):
        observed["config_enabled"] = config.enabled
        return fake_agent

    def _suggest_name(*, crop_path, agent, roster, ocr_raw=""):
        observed["crop_path"] = crop_path
        observed["agent"] = agent
        observed["roster"] = roster
        observed["ocr_raw"] = ocr_raw
        return ("陳小華", ["name.unconfirmed", "name.unconfirmed"])

    cfg = tmp_path / "name_agent.toml"
    cfg.write_text('enabled = true\nprovider = "claude"\n', encoding="utf-8")
    crop_path = tmp_path / "for testing only-page-0001-name.png"
    crop_path.write_bytes(b"fake png bytes")
    append_correction(
        tmp_path / "name_corrections.jsonl",
        Correction(field="name", final_value="葉心安", record_id="pdf-0001"),
    )
    pdf = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.pdf"
    fixture = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.ocr.json"
    out = tmp_path / "prepared.json"

    with pytest.MonkeyPatch.context() as monkeypatch:
        monkeypatch.setattr(
            "ocr_from2xlsx.prepare_records.prepare_records_from_paths",
            _prepare_records_from_paths,
        )
        monkeypatch.setattr("ocr_from2xlsx.name_agent.build_agent", _build_agent)
        monkeypatch.setattr("ocr_from2xlsx.name_suggestion.suggest_name", _suggest_name)
        code = main(
            [
                "prepare-records",
                "--input",
                str(pdf),
                "--output",
                str(out),
                "--ocr-fixture",
                str(fixture),
                "--name-agent-config",
                str(cfg),
            ]
        )

    assert code == 0
    data = _json.loads(out.read_text(encoding="utf-8"))
    assert data["records"][0]["name"] == "陳小華"
    assert data["records"][0]["ocr"]["warnings"] == ["existing-warning", "name.unconfirmed"]
    assert observed == {
        "config_enabled": True,
        "crop_path": str(crop_path),
        "agent": fake_agent,
        "roster": ["葉心安"],
        "ocr_raw": "raw",
    }


def test_resolve_name_crop_path_prefers_backend_supplied_name_crop_after_round_trip(tmp_path: Path) -> None:
    batch = Batch(
        source_batch=SourceBatch(
            created_at="2026-05-26T09:00:00+08:00",
            source_type="prepare_records",
            template_name="service_record.v1",
        ),
        records=[make_record("pdf-0001")],
    )
    record = batch.records[0]
    record.ocr.name_crop = "custom-crops/pdf-0001-handwritten-name.png"
    record.source.preprocessed_image_path = "fallback-page-0001.png"
    input_json = tmp_path / "prepared.json"
    expected_crop = tmp_path / "custom-crops" / "pdf-0001-handwritten-name.png"
    expected_crop.parent.mkdir(parents=True)
    expected_crop.write_bytes(b"fake png bytes")
    fallback_crop = tmp_path / "fallback-page-0001-name.png"
    fallback_crop.write_bytes(b"fallback bytes")

    dump_batch(batch, input_json)
    loaded = Batch.from_dict(json.loads(input_json.read_text(encoding="utf-8")))

    assert loaded.records[0].ocr.name_crop == "custom-crops/pdf-0001-handwritten-name.png"
    assert _resolve_name_crop_path(loaded.records[0], tmp_path) == str(expected_crop)


def test_resolve_name_crop_path_rejects_backend_path_outside_output_dir_and_uses_fallback(
    tmp_path: Path,
) -> None:
    record = make_record("pdf-0001")
    record.ocr.name_crop = "..\\escaped-name.png"
    record.source.preprocessed_image_path = "fallback-page-0001.png"
    fallback_crop = tmp_path / "fallback-page-0001-name.png"
    fallback_crop.write_bytes(b"fallback bytes")
    (tmp_path.parent / "escaped-name.png").write_bytes(b"backend bytes")

    assert _resolve_name_crop_path(record, tmp_path) == str(fallback_crop)


def test_resolve_name_crop_path_rejects_absolute_backend_path_outside_output_dir_and_uses_fallback(
    tmp_path: Path,
) -> None:
    record = make_record("pdf-0001")
    escaped_crop = tmp_path.parent / "absolute-escaped-name.png"
    escaped_crop.write_bytes(b"backend bytes")
    record.ocr.name_crop = str(escaped_crop.resolve())
    record.source.preprocessed_image_path = "fallback-page-0001.png"
    fallback_crop = tmp_path / "fallback-page-0001-name.png"
    fallback_crop.write_bytes(b"fallback bytes")

    assert _resolve_name_crop_path(record, tmp_path) == str(fallback_crop)


def test_prepare_records_enabled_name_agent_without_crop_is_noop(tmp_path):
    import json as _json
    from pathlib import Path as _Path

    from ocr_from2xlsx.cli import main

    record = make_record()
    record.name = ""
    record.ocr.warnings = []
    record.source.preprocessed_image_path = "for testing only-page-0001.png"

    def _prepare_records_from_paths(*args, **kwargs):
        return Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="prepare_records",
                template_name="service_record.v1",
            ),
            records=[record],
        )

    cfg = tmp_path / "name_agent.toml"
    cfg.write_text('enabled = true\nprovider = "claude"\n', encoding="utf-8")
    pdf = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.pdf"
    fixture = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.ocr.json"
    out = tmp_path / "prepared.json"

    with pytest.MonkeyPatch.context() as monkeypatch:
        monkeypatch.setattr(
            "ocr_from2xlsx.prepare_records.prepare_records_from_paths",
            _prepare_records_from_paths,
        )
        monkeypatch.setattr(
            "ocr_from2xlsx.name_suggestion.suggest_name",
            lambda **kwargs: pytest.fail("suggest_name should not be called without a crop"),
        )
        code = main(
            [
                "prepare-records",
                "--input",
                str(pdf),
                "--output",
                str(out),
                "--ocr-fixture",
                str(fixture),
                "--name-agent-config",
                str(cfg),
            ]
        )

    assert code == 0
    data = _json.loads(out.read_text(encoding="utf-8"))
    assert data["records"][0]["name"] == ""
    assert data["records"][0]["ocr"]["warnings"] == []


def test_prepare_records_enabled_name_agent_skips_malformed_correction_store_entries(tmp_path):
    import json as _json
    from pathlib import Path as _Path

    from ocr_from2xlsx.cli import main

    record = make_record()
    record.name = ""
    record.ocr.warnings = []
    record.source.preprocessed_image_path = "for testing only-page-0001.png"

    def _prepare_records_from_paths(*args, **kwargs):
        return Batch(
            source_batch=SourceBatch(
                created_at="2026-05-24T15:30:00+08:00",
                source_type="prepare_records",
                template_name="service_record.v1",
            ),
            records=[record],
        )

    class FakeAgent:
        pass

    fake_agent = FakeAgent()
    observed: dict[str, object] = {}

    def _build_agent(config):
        observed["config_enabled"] = config.enabled
        return fake_agent

    def _suggest_name(*, crop_path, agent, roster, ocr_raw=""):
        observed["crop_path"] = crop_path
        observed["agent"] = agent
        observed["roster"] = roster
        observed["ocr_raw"] = ocr_raw
        return ("陳小華", ["name.unconfirmed"])

    cfg = tmp_path / "name_agent.toml"
    cfg.write_text('enabled = true\nprovider = "claude"\n', encoding="utf-8")
    crop_path = tmp_path / "for testing only-page-0001-name.png"
    crop_path.write_bytes(b"fake png bytes")
    (tmp_path / "name_corrections.jsonl").write_text(
        "\n".join(
            [
                '{"field":"name","final_value":"葉心安","record_id":"pdf-0001"}',
                '{"field":"name","final_value":["bad"],"record_id":"pdf-0002"}',
                '{"field":"name","final_value":"broken"',
            ]
        ),
        encoding="utf-8",
    )
    pdf = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.pdf"
    fixture = _Path(__file__).parent / "fixtures" / "pdf" / "for testing only.ocr.json"
    out = tmp_path / "prepared.json"

    with pytest.MonkeyPatch.context() as monkeypatch:
        monkeypatch.setattr(
            "ocr_from2xlsx.prepare_records.prepare_records_from_paths",
            _prepare_records_from_paths,
        )
        monkeypatch.setattr("ocr_from2xlsx.name_agent.build_agent", _build_agent)
        monkeypatch.setattr("ocr_from2xlsx.name_suggestion.suggest_name", _suggest_name)
        code = main(
            [
                "prepare-records",
                "--input",
                str(pdf),
                "--output",
                str(out),
                "--ocr-fixture",
                str(fixture),
                "--name-agent-config",
                str(cfg),
            ]
        )

    assert code == 0
    data = _json.loads(out.read_text(encoding="utf-8"))
    assert data["records"][0]["name"] == "陳小華"
    assert data["records"][0]["ocr"]["warnings"] == ["name.unconfirmed"]
    assert observed == {
        "config_enabled": True,
        "crop_path": str(crop_path),
        "agent": fake_agent,
        "roster": ["葉心安"],
        "ocr_raw": "raw",
    }


def test_prepare_records_cli_reports_missing_ocr_fixture_page_without_traceback(
    tmp_path: Path, capsys
) -> None:
    pdf_path = tmp_path / "two-page.pdf"
    output_json = tmp_path / "prepared.json"
    missing_page_fixture = tmp_path / "missing-page.ocr.json"
    document = fitz.open()
    document.new_page(width=595.44, height=841.68)
    document.new_page(width=595.44, height=841.68)
    document.save(pdf_path)
    document.close()
    missing_page_fixture.write_text(
        json.dumps(
            {
                "pages": [
                    {
                        "document_name": pdf_path.name,
                        "page_number": 1,
                        "record": {
                            "record_id": "pdf-0001",
                            "service_date": "2026-05-26",
                            "identity": "patient",
                            "name": "AI test",
                            "medical_record_no": "TRAINING-ONLY",
                            "gender": "female",
                        },
                    }
                ]
            }
        ),
        encoding="utf-8",
    )

    exit_code = main(
        [
            "prepare-records",
            "--input",
            str(pdf_path),
            "--output",
            str(output_json),
            "--ocr-fixture",
            str(missing_page_fixture),
        ]
    )

    captured = capsys.readouterr()
    assert exit_code == 2
    assert captured.err.startswith("error: ")
    assert pdf_path.name in captured.err
    assert "page 2" in captured.err
    assert "Traceback" not in captured.err
