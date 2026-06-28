from __future__ import annotations

import sys
import tkinter as tk
from tkinter import ttk
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from ocr_from2xlsx import app as app_module
from ocr_from2xlsx import capture as capture_module
from ocr_from2xlsx.confirm_form import record_to_form_state
from ocr_from2xlsx.constants import BASIC_COLUMN_BY_FIELD
from ocr_from2xlsx.correction_store import load_corrections
from ocr_from2xlsx.app import ReviewApp
from ocr_from2xlsx.capture import CaptureResult
from ocr_from2xlsx.form_layout import service_record_layout
from ocr_from2xlsx.session import AcceptResult, ImportSession
from tests.fixtures import create_workbook_template
from tests.test_json_io import make_record


class StubSession:
    def __init__(self, result: AcceptResult) -> None:
        self.result = result
        self.calls: list[tuple[str, bool, bool]] = []
        self.overwrite_rows: list[int | None] = []

    def accept_scan(
        self,
        record,
        force: bool = False,
        human_confirmed: bool = False,
        overwrite_row: int | None = None,
        relaxed: bool = False,
    ) -> AcceptResult:
        self.calls.append((record.record_id, force, human_confirmed))
        self.overwrite_rows.append(overwrite_row)
        if human_confirmed and self.result.status in {"forced", "written"}:
            record.ocr.warnings = [warning for warning in record.ocr.warnings if warning != "name.unconfirmed"]
        return self.result

    def close(self) -> None:
        return None


class FakeVar:
    def __init__(self, value: str = "") -> None:
        self._value = value

    def set(self, value: str) -> None:
        self._value = value

    def get(self) -> str:
        return self._value


class FakeListbox:
    def __init__(self) -> None:
        self.items: list[str] = []

    def insert(self, _index, message: str) -> None:
        self.items.append(message)

    def see(self, _index) -> None:
        return None


class FakePreview:
    # Mirrors the ImageViewer interface (#47): show_image (static/pannable),
    # show_frame (live), show_placeholder, frame_region. Tracks .image/.text/.mode
    # for the camera/preview assertions.
    def __init__(self) -> None:
        self.text = ""
        self.image = None
        self.mode = "placeholder"
        self.reset_called = False

    def reset_view(self) -> None:
        self.reset_called = True

    def show_image(self, image) -> None:
        self.mode = "static"
        self.image = image

    def show_frame(self, image) -> None:
        self.mode = "live"
        self.image = image

    def show_placeholder(self, text: str) -> None:
        self.mode = "placeholder"
        self.image = None
        self.text = text

    def frame_region(self, band) -> None:
        return None

    def get(self, _start: str, _end: str) -> str:
        return self.text


class FakeConfirmForm:
    def __init__(self, fields: dict[str, FakeVar]) -> None:
        self._fields = fields
        self.state: dict[str, object] = {}

    def prefill(self, state: dict[str, object]) -> None:
        self.state = dict(state)
        for key in ("service_date", "identity", "name", "medical_record_no", "gender"):
            if key in self._fields and key in state:
                value = state[key]
                self._fields[key].set("" if value is None else str(value))

    def set_flagged_fields(self, flagged: dict[str, str]) -> None:
        self.flagged = dict(flagged)

    def collect(self) -> dict[str, object]:
        collected = dict(self.state)
        for key in ("service_date", "identity", "name", "medical_record_no", "gender"):
            if key in self._fields:
                collected[key] = self._fields[key].get()
        return collected

    def flagged_keys(self) -> list[str]:
        # NOTE: returns dict insertion order, not the real ConfirmForm's layout order.
        # Fine for the headless tests here (single flag); assert multi-flag focus order
        # against the real ConfirmForm (tests/test_confirm_form_keyboard.py) instead.
        return list(getattr(self, "flagged", {}).keys())

    def flagged_count(self) -> int:
        return len(getattr(self, "flagged", {}))

    def focus_first_flagged(self) -> str | None:
        keys = self.flagged_keys()
        self.focused = keys[0] if keys else None
        return self.focused

    def focus_next_flagged(self) -> str | None:
        return None

    def focus_prev_flagged(self) -> str | None:
        return None


class _FakePreviewCapture:
    def __init__(
        self,
        *,
        opened: bool,
        frames: list[object] | None = None,
        failed_reads_before_frame: int = 0,
    ) -> None:
        self._opened = opened
        self._frames = list(frames or [])
        self._failed_reads_before_frame = failed_reads_before_frame
        self.released = False

    def isOpened(self) -> bool:
        return self._opened

    def release(self) -> None:
        self.released = True

    def read(self) -> tuple[bool, object | None]:
        if self._failed_reads_before_frame > 0:
            self._failed_reads_before_frame -= 1
            return False, None
        if self._frames:
            return True, self._frames.pop(0)
        return False, None


def _column_for_header(sheet, header: str) -> int:
    for cell in sheet[1]:
        if cell.value == header:
            return cell.column
    raise AssertionError(f"Missing header in fixture: {header}")


def _button_texts(widget: tk.Misc) -> list[str]:
    texts: list[str] = []
    for child in widget.winfo_children():
        if isinstance(child, ttk.Button):
            texts.append(child.cget("text"))
        texts.extend(_button_texts(child))
    return texts


def _preview_text(preview: object) -> str:
    if isinstance(preview, tk.Text):
        return preview.get("1.0", "end-1c")
    if hasattr(preview, "get"):
        return preview.get("1.0", "end-1c")
    return str(preview)


def test_review_app_builds_confirm_form_and_prefills_record(tmp_path: Path) -> None:
    try:
        app = ReviewApp()
    except tk.TclError as exc:
        pytest.skip(f"no display available for Tk: {exc}")

    app.withdraw()
    try:
        record = make_record("scan-0003")
        app.loaded_json_path = tmp_path / "records.json"

        app._show_record(record)

        expected_state = record_to_form_state(app.layout, record)
        collected = app.confirm_form.collect()
        assert app.layout == service_record_layout()
        assert collected["service_date"] == expected_state["service_date"]
        assert collected["identity"] == expected_state["identity"]
        assert collected["name"] == expected_state["name"]
        assert collected["medical_record_no"] == expected_state["medical_record_no"]
        assert collected["gender"] == expected_state["gender"]
        assert collected["cancer"] == expected_state["cancer"]
        # Slim toolbar keeps only the most-used actions; the rest moved to the menu bar (#56).
        toolbar_texts = _button_texts(app)
        assert {"開啟報表", "匯入資料夾", "上一筆", "下一筆", "確認並寫入"} <= set(toolbar_texts)
        # #remove-add-page-button: the manual 新增頁面 button is gone from the toolbar — the flow
        # auto-opens a fresh page on 開啟報表 and after 確認並寫入. It stays in the 編輯 menu only.
        assert "新增頁面" not in toolbar_texts
    finally:
        app.destroy()


def test_menu_bar_categories_and_state_machine_drives_menu_entries() -> None:
    try:
        app = ReviewApp()
    except tk.TclError as exc:
        pytest.skip(f"no display available for Tk: {exc}")
    app.withdraw()
    try:
        menubar = app.nametowidget(app.cget("menu"))
        labels = {menubar.entrycget(i, "label") for i in range(menubar.index("end") + 1)}
        assert {"檔案(F)", "掃描(S)", "編輯(E)", "檢視(V)", "說明(H)"} <= labels

        # The state machine (#56) gates MENU entries, not just toolbar buttons: with no records
        # 上一筆 (a menu entry) is disabled. 確認並寫入 / 強制寫入 are deliberately NOT gated —
        # they stay clickable so they can surface a clear "缺少工作檔"/"姓名未填" error instead
        # of greying out silently (#confirm-required-fields).
        app.session = None
        app.records = []
        app.current_index = -1
        app._update_toolbar_states()
        edit_menu = app.nametowidget(menubar.entrycget("編輯(E)", "menu"))
        assert str(edit_menu.entrycget("上一筆", "state")) == "disabled"
        assert str(edit_menu.entrycget("確認並寫入", "state")) == "normal"
        assert str(edit_menu.entrycget("強制寫入", "state")) == "normal"
    finally:
        app.destroy()


def test_confirm_current_writes_whole_page_and_clears_unconfirmed_name(tmp_path: Path) -> None:
    try:
        app = ReviewApp()
    except tk.TclError as exc:
        pytest.skip(f"no display available for Tk: {exc}")

    app.withdraw()
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    session = ImportSession.start(template, working)
    try:
        first = make_record("scan-0001")
        first.ocr.warnings = ["name.unconfirmed"]
        first.ocr.raw_text = "癌症資源中心服務紀錄表\n姓名 王小明"
        second = make_record("scan-0002")
        app.session = session
        app.records = [first, second]
        app.loaded_json_path = tmp_path / "records.json"
        app.correction_store_path = tmp_path / "name_corrections.jsonl"

        app._next_record()
        app.confirm_form.text_fields["name"].set(" 王小明 ")
        app.confirm_form.multi_choice_fields["cancer"]["lung_cancer"].set(True)

        app._confirm_current()

        assert app.written_indices == {0}
        assert app.current_index == 1
        assert app.fields["record_id"].get() == second.record_id
        assert first.name == "王小明"
        assert first.ocr.warnings == []
        assert set(first.patient_fields.cancers) == {"breast_cancer", "lung_cancer"}
        assert load_corrections(app.correction_store_path)[0].final_value == "王小明"
        assert any("scan-0001: written" in status for status in app._status_log)

        workbook = load_workbook(working)
        try:
            sheet = workbook["個案總表"]
            name_col = _column_for_header(sheet, BASIC_COLUMN_BY_FIELD["name"])
            assert sheet.cell(row=2, column=name_col).value == "王小明"
        finally:
            workbook.close()
    finally:
        session.close()
        app.destroy()


def test_show_record_without_source_image_keeps_placeholder_preview(tmp_path: Path) -> None:
    try:
        app = ReviewApp()
    except tk.TclError as exc:
        pytest.skip(f"no display available for Tk: {exc}")

    app.withdraw()
    try:
        record = make_record("scan-0004")
        record.source.preprocessed_image_path = "missing-preview.png"
        app.loaded_json_path = tmp_path / "records.json"

        app._show_record(record)

        assert "攝影機或圖片預覽區" in _preview_text(app.preview)
    finally:
        app.destroy()


def test_confirm_form_round_trips_prefilled_state() -> None:
    try:
        root = tk.Tk()
    except tk.TclError as exc:
        pytest.skip(f"no display available for Tk: {exc}")

    root.withdraw()
    try:
        form = app_module.ConfirmForm(root, service_record_layout())
        state = {
            "service_date": "2026-05-31",
            "identity": "patient",
            "name": "王小明",
            "medical_record_no": "A123456",
            "diagnosis_date": "2026-01-15",
            "gender": "female",
            "nationality": "local",
            "age": "51_60",
            "channel": "internal_referral",
            "disease_status": "treating",
            "source": "outpatient",
            "newly_diagnosed": "true",
            "consultation.health_medical": {"screening_prevention", "other"},
            "consultation.care_support": {"peer_experience"},
            "supplies": {"wig_hat"},
            "cancer": {"breast_cancer", "lung_cancer"},
        }

        form.prefill(state)

        assert isinstance(form.frame, tk.Widget)
        assert isinstance(form.text_fields["service_date"], tk.StringVar)
        assert isinstance(form.single_choice_fields["identity"], tk.StringVar)
        assert isinstance(form.multi_choice_fields["cancer"]["breast_cancer"], tk.BooleanVar)
        collected = form.collect()
        for key, value in state.items():
            assert collected[key] == value
    finally:
        root.destroy()


def test_confirm_form_single_choice_can_clear_to_unselected() -> None:
    try:
        root = tk.Tk()
    except tk.TclError as exc:
        pytest.skip(f"no display available for Tk: {exc}")

    root.withdraw()
    try:
        form = app_module.ConfirmForm(root, service_record_layout())

        form.prefill({"nationality": "local"})
        assert form.collect()["nationality"] == "local"

        # No "清除" button anymore (#31): single choice is checkboxes that clear by
        # toggling the selected one off; the field can return to unselected.
        assert not hasattr(form, "single_choice_clear_buttons")
        form.single_choice_fields["nationality"].set("")

        assert form.collect()["nationality"] == ""
    finally:
        root.destroy()


def test_single_choice_checkboxes_reflect_selection() -> None:
    try:
        root = tk.Tk()
    except tk.TclError as exc:
        pytest.skip(f"no display available for Tk: {exc}")

    root.withdraw()
    try:
        form = app_module.ConfirmForm(root, service_record_layout())
        form.prefill({"identity": "patient"})
        option_vars = form._single_choice_option_vars["identity"]
        assert option_vars["patient"].get() is True
        assert all(not var.get() for code, var in option_vars.items() if code != "patient")
    finally:
        root.destroy()


@pytest.fixture
def app(monkeypatch: pytest.MonkeyPatch) -> ReviewApp:
    monkeypatch.setattr("ocr_from2xlsx.app.messagebox.showerror", lambda *args, **kwargs: None)
    monkeypatch.setattr("ocr_from2xlsx.app.messagebox.showinfo", lambda *args, **kwargs: None)
    monkeypatch.setattr("ocr_from2xlsx.app.messagebox.showwarning", lambda *args, **kwargs: None)
    review_app = ReviewApp.__new__(ReviewApp)
    review_app.records = []
    review_app.current_index = -1
    review_app.session = None
    review_app.editing = False
    review_app.written_indices = set()
    review_app._written_rows = {}
    review_app._blocked_indices = set()
    review_app.loaded_json_path = None
    review_app.correction_store_path = None
    review_app.layout = service_record_layout()
    review_app.fields = {
        "record_id": FakeVar(),
        "service_date": FakeVar(),
        "identity": FakeVar(),
        "name": FakeVar(),
        "medical_record_no": FakeVar(),
        "gender": FakeVar(),
    }
    review_app.confirm_form = FakeConfirmForm(review_app.fields)
    review_app.preview = FakePreview()
    review_app._preview_image = None
    review_app._preview_rotation = 0
    review_app._status_log = []
    review_app._status_var = None
    review_app._pending_var = None
    review_app._progress_var = None
    review_app._badge_var = None
    review_app._status_log_path = None
    review_app._recognition_threaded = False  # run recognition inline in the headless harness
    return review_app


def test_force_write_does_not_advance(app: ReviewApp) -> None:
    records = [make_record("scan-0001"), make_record("scan-0002")]
    app.records = records
    app.current_index = 0
    app._show_record(records[0])
    app.editing = True
    result = AcceptResult(
        record_id=records[0].record_id,
        status="forced",
        row_number=2,
        blockers=["patient.nationality.required"],
        warnings=[],
    )
    session = StubSession(result)
    app.session = session

    app._force_write()

    assert app.current_index == 0
    assert app.editing is False
    assert app.written_indices == {0}
    assert session.calls == [(records[0].record_id, True, True)]
    assert app.fields["record_id"].get() == records[0].record_id


def test_force_write_persists_non_legacy_confirm_form_fields(app: ReviewApp) -> None:
    record = make_record("scan-0005")
    app.records = [record]
    app.current_index = 0
    app._show_record(record)
    app.editing = True
    app.confirm_form.state["cancer"] = {"lung_cancer"}
    result = AcceptResult(
        record_id=record.record_id,
        status="forced",
        row_number=2,
        blockers=[],
        warnings=[],
    )
    app.session = StubSession(result)

    app._force_write()

    assert record.patient_fields.cancers == ["lung_cancer"]
    assert record.review.edited_by_user is True


def test_force_write_skips_when_already_written(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    records = [make_record("scan-0001")]
    app.records = records
    app.current_index = 0
    app._show_record(records[0])
    app.written_indices = {0}
    info_calls: list[tuple[str, str]] = []

    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showinfo",
        lambda title, message: info_calls.append((title, message)),
    )

    class FailingSession:
        def accept_scan(self, record, force: bool = False, overwrite_row: int | None = None) -> AcceptResult:
            raise AssertionError("accept_scan should not be called")

        def close(self) -> None:
            return None

    app.session = FailingSession()

    app._force_write()

    assert info_calls == [("提示", "目前資料已寫入，請切換下一筆。")]
    assert app.written_indices == {0}


def test_next_record_browses_without_writing_current_record(app: ReviewApp) -> None:
    records = [make_record("scan-0001"), make_record("scan-0002")]
    app.records = records
    app.current_index = 0
    app._show_record(records[0])

    class FailingSession:
        def accept_scan(self, record, force: bool = False, overwrite_row: int | None = None) -> AcceptResult:
            raise AssertionError("accept_scan should not be called")

        def close(self) -> None:
            return None

    app.session = FailingSession()

    app._next_record()

    assert app.current_index == 1
    assert app.fields["record_id"].get() == records[1].record_id
    assert app.written_indices == set()


def test_previous_record_browses_without_writing_current_record(app: ReviewApp) -> None:
    records = [make_record("scan-0001"), make_record("scan-0002")]
    app.records = records
    app.current_index = 1
    app._show_record(records[1])

    class FailingSession:
        def accept_scan(self, record, force: bool = False, overwrite_row: int | None = None) -> AcceptResult:
            raise AssertionError("accept_scan should not be called")

        def close(self) -> None:
            return None

    app.session = FailingSession()

    ReviewApp._previous_record(app)

    assert app.current_index == 0
    assert app.fields["record_id"].get() == records[0].record_id


def test_next_record_from_initial_index_shows_first_record(app: ReviewApp) -> None:
    record = make_record("scan-0001")
    app.records = [record]

    class FailingSession:
        def accept_scan(self, record, force: bool = False, overwrite_row: int | None = None) -> AcceptResult:
            raise AssertionError("accept_scan should not be called")

        def close(self) -> None:
            return None

    app.session = FailingSession()

    app._next_record()

    assert app.current_index == 0
    assert app.fields["record_id"].get() == record.record_id


def test_load_json_sets_default_correction_store_path(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    json_path = tmp_path / "records.json"
    records = [make_record("scan-0001")]

    monkeypatch.setattr(
        "ocr_from2xlsx.app.filedialog.askopenfilename", lambda **kwargs: str(json_path)
    )

    class FakeSource:
        def __init__(self, path: str) -> None:
            assert Path(path) == json_path

        def records(self):
            return iter(records)

    monkeypatch.setattr("ocr_from2xlsx.app.JsonRecordSource", FakeSource)

    app._load_json()

    assert app.loaded_json_path == json_path
    assert app.correction_store_path == json_path.parent / "name_corrections.jsonl"


def test_start_camera_falls_back_to_directshow_and_keeps_selected_index(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    calls: list[tuple[int, ...]] = []
    stop_calls: list[str] = []
    statuses: list[str] = []
    placeholders: list[str] = []
    plain_capture = _FakePreviewCapture(opened=False)
    directshow_capture = _FakePreviewCapture(opened=True, frames=["frame"])

    def fake_video_capture(index: int, backend: int | None = None) -> _FakePreviewCapture:
        calls.append((index,) if backend is None else (index, backend))
        if backend is None:
            return plain_capture
        return directshow_capture

    monkeypatch.setitem(
        sys.modules,
        "cv2",
        SimpleNamespace(CAP_DSHOW=700, VideoCapture=fake_video_capture),
    )
    monkeypatch.setattr(app, "_stop_camera", lambda: stop_calls.append("stop"))
    monkeypatch.setattr(app, "_push_status", lambda message: statuses.append(message))
    monkeypatch.setattr(app, "_show_placeholder_preview", lambda: placeholders.append("placeholder"))
    monkeypatch.setattr(app, "after", lambda delay, callback: f"after-{delay}")
    app._camera_capture = None
    app._camera_after_id = None
    app._camera_failure_count = 3
    app._camera_index = None

    ReviewApp._start_camera(app, 4)

    assert stop_calls == ["stop"]
    assert calls == [(4,), (4, 700)]
    assert plain_capture.released is True
    assert directshow_capture.released is False
    assert app._camera_capture is directshow_capture
    assert app._camera_after_id == "after-0"
    assert app._camera_failure_count == 0
    assert app._camera_index == 4
    assert statuses == ["攝影機已連接（裝置 #4）"]
    assert placeholders == []


def test_start_camera_skips_backend_that_opens_but_cannot_read_frames(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    calls: list[tuple[int, ...]] = []
    stop_calls: list[str] = []
    statuses: list[str] = []
    placeholders: list[str] = []
    plain_capture = _FakePreviewCapture(opened=True)
    directshow_capture = _FakePreviewCapture(opened=True, frames=["frame"])

    def fake_video_capture(index: int, backend: int | None = None) -> _FakePreviewCapture:
        calls.append((index,) if backend is None else (index, backend))
        if backend is None:
            return plain_capture
        return directshow_capture

    monkeypatch.setitem(
        sys.modules,
        "cv2",
        SimpleNamespace(CAP_DSHOW=700, VideoCapture=fake_video_capture),
    )
    monkeypatch.setattr(app, "_stop_camera", lambda: stop_calls.append("stop"))
    monkeypatch.setattr(app, "_push_status", lambda message: statuses.append(message))
    monkeypatch.setattr(app, "_show_placeholder_preview", lambda: placeholders.append("placeholder"))
    monkeypatch.setattr(app, "after", lambda delay, callback: f"after-{delay}")
    app._camera_capture = None
    app._camera_after_id = None
    app._camera_failure_count = 3
    app._camera_index = None

    ReviewApp._start_camera(app, 4)

    assert stop_calls == ["stop"]
    assert calls == [(4,), (4, 700)]
    assert plain_capture.released is True
    assert directshow_capture.released is False
    assert app._camera_capture is directshow_capture
    assert app._camera_after_id == "after-0"
    assert app._camera_failure_count == 0
    assert app._camera_index == 4
    assert statuses == ["攝影機已連接（裝置 #4）"]
    assert placeholders == []


def test_start_camera_accepts_slow_start_plain_backend(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    calls: list[tuple[int, ...]] = []
    stop_calls: list[str] = []
    statuses: list[str] = []
    placeholders: list[str] = []
    first_capture = _FakePreviewCapture(
        opened=True,
        frames=["frame"],
        failed_reads_before_frame=capture_module.DEFAULT_CAMERA_PROBE_READS + 1,
    )
    second_capture = _FakePreviewCapture(
        opened=True,
        frames=["frame"],
        failed_reads_before_frame=capture_module.DEFAULT_CAMERA_PROBE_READS + 1,
    )
    captures = iter([first_capture, second_capture])

    def fake_video_capture(index: int, backend: int | None = None) -> _FakePreviewCapture:
        calls.append((index,) if backend is None else (index, backend))
        return next(captures)

    monkeypatch.setitem(
        sys.modules,
        "cv2",
        SimpleNamespace(VideoCapture=fake_video_capture),
    )
    monkeypatch.setattr(app, "_stop_camera", lambda: stop_calls.append("stop"))
    monkeypatch.setattr(app, "_push_status", lambda message: statuses.append(message))
    monkeypatch.setattr(app, "_show_placeholder_preview", lambda: placeholders.append("placeholder"))
    monkeypatch.setattr(app, "after", lambda delay, callback: f"after-{delay}")
    app._camera_capture = None
    app._camera_after_id = None
    app._camera_failure_count = 3
    app._camera_index = None

    ReviewApp._start_camera(app, 4)

    assert stop_calls == ["stop"]
    assert calls == [(4,), (4,)]
    assert first_capture.released is True
    assert second_capture.released is False
    assert app._camera_capture is second_capture
    assert app._camera_after_id == "after-0"
    assert app._camera_failure_count == 0
    assert app._camera_index == 4
    assert statuses == ["攝影機已連接（裝置 #4）"]
    assert placeholders == []


def test_start_camera_clears_stale_index_when_startup_fails(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    stop_calls: list[str] = []
    statuses: list[str] = []
    placeholders: list[str] = []
    monkeypatch.setattr("ocr_from2xlsx.app.open_camera_capture", lambda index: None)
    monkeypatch.setattr(app, "_stop_camera", lambda: stop_calls.append("stop"))
    monkeypatch.setattr(app, "_push_status", lambda message: statuses.append(message))
    monkeypatch.setattr(app, "_show_placeholder_preview", lambda: placeholders.append("placeholder"))
    app._camera_capture = None
    app._camera_after_id = None
    app._camera_failure_count = 1
    app._camera_index = 7

    ReviewApp._start_camera(app, 4)

    assert stop_calls == ["stop"]
    assert statuses == ["無法開啟攝影機 4"]
    assert placeholders == ["placeholder"]
    assert app._camera_capture is None
    assert app._camera_after_id is None
    assert app._camera_failure_count == 1
    assert app._camera_index is None


def test_choose_camera_cancel_clears_stale_index_without_live_preview(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    start_calls: list[int] = []
    app._camera_capture = None
    app._camera_after_id = None
    app._camera_index = 7
    monkeypatch.setattr(app, "_ask_camera", lambda indices: None)
    monkeypatch.setattr(app, "_start_camera", lambda index: start_calls.append(index))

    ReviewApp._choose_camera(app, [4, 7])

    assert start_calls == []
    assert app._camera_index is None


def test_capture_and_recognize_warns_when_camera_is_missing_without_live_preview(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    import ocr_from2xlsx.capture as capture_module

    warnings: list[tuple[str, str]] = []
    capture_calls: list[tuple[tuple[object, ...], dict[str, object]]] = []
    app._camera_capture = None
    app._camera_after_id = None
    app._camera_index = None

    def fail_capture(*args, **kwargs):
        capture_calls.append((args, kwargs))
        raise AssertionError("capture_still should not be called without a selected camera")

    monkeypatch.setattr(capture_module, "require_camera_support", lambda: None)
    monkeypatch.setattr(capture_module, "capture_still", fail_capture)
    monkeypatch.setattr(
        app,
        "_recognize_capture",
        lambda frame: (_ for _ in ()).throw(AssertionError("should not recognize")),
    )
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showwarning",
        lambda title, message: warnings.append((title, message)),
    )

    ReviewApp._capture_and_recognize(app)

    assert warnings == [("擷取並辨識", "請先選擇可用的攝影機。")]
    assert capture_calls == []


def test_capture_and_recognize_reports_missing_opencv_before_camera_selection_warning(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    import builtins
    import ocr_from2xlsx.capture as capture_module
    import sys

    messages: list[tuple[str, str, str]] = []
    capture_calls: list[tuple[tuple[object, ...], dict[str, object]]] = []
    original_import = builtins.__import__
    app._camera_capture = None
    app._camera_after_id = None
    app._camera_index = None

    def fail_capture(*args, **kwargs):
        capture_calls.append((args, kwargs))
        raise AssertionError("capture_still should not be called without a selected camera")

    def no_cv2(
        name: str,
        globals: object | None = None,
        locals: object | None = None,
        fromlist: tuple[str, ...] = (),
        level: int = 0,
    ) -> object:
        if name == "cv2":
            raise ImportError("no cv2")
        return original_import(name, globals, locals, fromlist, level)

    monkeypatch.delitem(sys.modules, "cv2", raising=False)
    monkeypatch.setattr(builtins, "__import__", no_cv2)
    monkeypatch.setattr(capture_module, "capture_still", fail_capture)
    monkeypatch.setattr(
        app,
        "_recognize_capture",
        lambda frame: (_ for _ in ()).throw(AssertionError("should not recognize")),
    )
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showwarning",
        lambda title, message: messages.append(("warning", title, message)),
    )
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showerror",
        lambda title, message: messages.append(("error", title, message)),
    )

    ReviewApp._capture_and_recognize(app)

    assert capture_calls == []
    assert len(messages) == 1
    assert messages[0][0] == "error"
    assert messages[0][1] == "擷取並辨識"
    assert "OpenCV" in messages[0][2]
    assert "pip install" in messages[0][2]


def test_capture_and_recognize_blocks_when_current_record_has_unsaved_edits(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    import ocr_from2xlsx.capture as capture_module

    record = make_record("scan-0001")
    errors: list[tuple[str, str]] = []
    stop_calls: list[str] = []
    capture_calls: list[tuple[tuple[object, ...], dict[str, object]]] = []
    app.records = [record]
    app.current_index = 0
    app.loaded_json_path = Path("C:\\scan-output\\scan-prepared.json")
    app._show_record(record)
    app.editing = True
    app._camera_capture = object()
    app._camera_after_id = "after-33"
    app._camera_index = 7

    def fail_capture(*args, **kwargs):
        capture_calls.append((args, kwargs))
        raise AssertionError("capture_still should not run while edits are unsaved")

    monkeypatch.setattr(capture_module, "capture_still", fail_capture)
    monkeypatch.setattr(app, "_stop_camera", lambda: stop_calls.append("stop"))
    monkeypatch.setattr(
        app,
        "_recognize_capture",
        lambda frame: (_ for _ in ()).throw(AssertionError("should not recognize")),
    )
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showerror",
        lambda title, message: errors.append((title, message)),
    )

    ReviewApp._capture_and_recognize(app)

    assert errors == [("尚未保存", "目前資料已修改，請先使用「確認並寫入」或「強制寫入」。")]
    assert stop_calls == []
    assert capture_calls == []
    assert app.records == [record]
    assert app.current_index == 0
    assert app.fields["record_id"].get() == record.record_id
    assert app.loaded_json_path == Path("C:\\scan-output\\scan-prepared.json")


def test_capture_and_recognize_warns_when_capture_is_blurry_without_live_preview(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    import ocr_from2xlsx.capture as capture_module

    warnings: list[tuple[str, str]] = []
    calls: list[str] = []
    app._camera_capture = None
    app._camera_after_id = None
    app._camera_index = 4
    monkeypatch.setattr(
        capture_module,
        "capture_still",
        lambda *args, **kwargs: CaptureResult(
            frame=object(),
            resolution=(1920, 1080),
            sharpness=12.4,
            brightness=128.0,
            passed=False,
        ),
    )
    monkeypatch.setattr(app, "_stop_camera", lambda: calls.append("stop"))
    monkeypatch.setattr(app, "_init_camera", lambda: calls.append("init"))
    monkeypatch.setattr(
        app,
        "_recognize_capture",
        lambda frame: (_ for _ in ()).throw(AssertionError("should not recognize")),
    )
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showwarning",
        lambda title, message: warnings.append((title, message)),
    )

    ReviewApp._capture_and_recognize(app)

    assert warnings == [
        ("擷取並辨識", "畫面太模糊（清晰度 12）。請調整對焦/光線/距離後重試。")
    ]
    assert calls == ["stop"]


def test_capture_and_recognize_delegates_without_restore_when_preview_not_live(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    # Recognition (and its error/restore handling) now lives in _recognize_capture, which runs
    # off-thread; _capture_and_recognize captures, then delegates with the restore info.
    import ocr_from2xlsx.capture as capture_module

    captured: dict = {}
    calls: list[str] = []
    start_calls: list[int] = []
    app._camera_capture = None
    app._camera_after_id = None
    app._camera_index = 7
    monkeypatch.setattr(
        capture_module,
        "capture_still",
        lambda *args, **kwargs: CaptureResult(
            frame="frame", resolution=(1920, 1080), sharpness=180.0, brightness=128.0, passed=True
        ),
    )
    monkeypatch.setattr(app, "_stop_camera", lambda: calls.append("stop"))
    monkeypatch.setattr(app, "_start_camera", lambda index: start_calls.append(index))
    monkeypatch.setattr(app, "_recognize_capture", lambda frame, **kw: captured.update(frame=frame, kw=kw))

    ReviewApp._capture_and_recognize(app)

    assert captured["frame"] == "frame"
    assert captured["kw"] == {"restore_live_preview": False, "restore_index": 7}
    assert calls == ["stop"]
    assert start_calls == []  # capture itself does not restart the camera; recognise handles restore


def test_capture_and_recognize_delegates_with_restore_when_preview_was_live(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    import ocr_from2xlsx.capture as capture_module

    captured: dict = {}
    app._camera_capture = object()
    app._camera_after_id = "after-33"
    app._camera_index = 7
    monkeypatch.setattr(
        capture_module,
        "capture_still",
        lambda *args, **kwargs: CaptureResult(
            frame="frame", resolution=(1920, 1080), sharpness=180.0, brightness=128.0, passed=True
        ),
    )
    monkeypatch.setattr(app, "_stop_camera", lambda: None)
    monkeypatch.setattr(app, "_start_camera", lambda index: None)
    monkeypatch.setattr(app, "_recognize_capture", lambda frame, **kw: captured.update(kw=kw))

    ReviewApp._capture_and_recognize(app)

    # A live preview must be handed to _recognize_capture so it can restore it on abort.
    assert captured["kw"] == {"restore_live_preview": True, "restore_index": 7}


def test_capture_and_recognize_does_not_reopen_preview_after_no_camera_warning(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    import ocr_from2xlsx.capture as capture_module

    warnings: list[tuple[str, str]] = []
    stop_calls: list[str] = []
    start_calls: list[int] = []
    app._camera_capture = object()
    app._camera_after_id = "after-33"
    app._camera_index = 7

    def stop_camera() -> None:
        stop_calls.append("stop")
        app._camera_capture = None
        app._camera_after_id = None

    monkeypatch.setattr(capture_module, "capture_still", lambda *args, **kwargs: None)
    monkeypatch.setattr(app, "_stop_camera", stop_camera)
    monkeypatch.setattr(app, "_start_camera", lambda index: start_calls.append(index))
    monkeypatch.setattr(
        app,
        "_recognize_capture",
        lambda frame: (_ for _ in ()).throw(AssertionError("should not recognize")),
    )
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showwarning",
        lambda title, message: warnings.append((title, message)),
    )

    ReviewApp._capture_and_recognize(app)

    assert warnings == [("擷取並辨識", "找不到可用的攝影機。")]
    assert stop_calls == ["stop"]
    assert start_calls == []
    assert app._camera_index is None


def test_capture_and_recognize_reports_missing_opencv_with_install_guidance(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    import builtins
    import sys

    messages: list[tuple[str, str, str]] = []
    stop_calls: list[str] = []
    start_calls: list[int] = []
    original_import = builtins.__import__
    app._camera_capture = object()
    app._camera_after_id = "after-33"
    app._camera_index = 7

    def stop_camera() -> None:
        stop_calls.append("stop")
        app._camera_capture = None
        app._camera_after_id = None

    def no_cv2(
        name: str,
        globals: object | None = None,
        locals: object | None = None,
        fromlist: tuple[str, ...] = (),
        level: int = 0,
    ) -> object:
        if name == "cv2":
            raise ImportError("no cv2")
        return original_import(name, globals, locals, fromlist, level)

    monkeypatch.delitem(sys.modules, "cv2", raising=False)
    monkeypatch.setattr(builtins, "__import__", no_cv2)
    monkeypatch.setattr(app, "_stop_camera", stop_camera)
    monkeypatch.setattr(app, "_start_camera", lambda index: start_calls.append(index))
    monkeypatch.setattr(
        app,
        "_recognize_capture",
        lambda frame: (_ for _ in ()).throw(AssertionError("should not recognize")),
    )
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showwarning",
        lambda title, message: messages.append(("warning", title, message)),
    )
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showerror",
        lambda title, message: messages.append(("error", title, message)),
    )

    ReviewApp._capture_and_recognize(app)

    assert len(messages) == 1
    assert messages[0][1] == "擷取並辨識"
    assert "OpenCV" in messages[0][2]
    assert "pip install" in messages[0][2]
    assert "找不到可用的攝影機" not in messages[0][2]
    assert stop_calls == ["stop"]
    assert start_calls == []
    assert app._camera_index is None


def test_capture_and_recognize_keeps_recognized_still_preview_on_success(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    import ocr_from2xlsx.capture as capture_module

    start_calls: list[int] = []
    app._camera_index = 7
    monkeypatch.setattr(
        capture_module,
        "capture_still",
        lambda *args, **kwargs: CaptureResult(
            frame="frame",
            resolution=(1920, 1080),
            sharpness=180.0,
            brightness=128.0,
            passed=True,
        ),
    )
    monkeypatch.setattr(app, "_stop_camera", lambda: None)
    monkeypatch.setattr(app, "_start_camera", lambda index: start_calls.append(index))
    monkeypatch.setattr(
        app,
        "_recognize_capture",
        lambda frame, **kw: setattr(app.preview, "image", "recognized-still"),
    )

    ReviewApp._capture_and_recognize(app)

    assert start_calls == []
    assert app.preview.image == "recognized-still"


def test_recognize_capture_writes_json_and_loads_review_flow(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    output_dir = tmp_path / "scan-output"

    class _FakeBackend:
        def extract(self, prepared) -> dict[str, object]:
            return {
                "service_date": "2025-06-25",
                "identity": "patient",
                "gender": "female",
                "ocr": {
                    "backend": "fake",
                    "raw_text": str(prepared.image_path),
                    "warnings": [],
                },
            }

    def fake_imwrite(path: str, frame: object) -> bool:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        Path(path).write_bytes(b"\x89PNG\r\n")
        return True

    monkeypatch.setattr(
        "ocr_from2xlsx.app.filedialog.askdirectory", lambda **kwargs: str(output_dir)
    )
    monkeypatch.setenv("OCR_BACKEND", "plugin")
    monkeypatch.setattr(
        "ocr_from2xlsx.plugin_backend.PluginOcrBackend.resolve",
        lambda explicit_dir=None, default_dir=None: _FakeBackend(),
    )
    monkeypatch.setitem(sys.modules, "cv2", SimpleNamespace(imwrite=fake_imwrite))

    ReviewApp._recognize_capture(app, frame=object())

    assert app.loaded_json_path == output_dir / "scan-prepared.json"
    assert app.correction_store_path == output_dir / "name_corrections.jsonl"
    assert app.current_index == 0
    assert app.fields["record_id"].get() == "scan-0001"
    assert (output_dir / "scan-capture.png").is_file()
    assert (output_dir / "scan-prepared.json").is_file()


def test_recognize_capture_explicitly_passes_scan_docpre_opt_in(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    output_dir = tmp_path / "scan-output"
    calls: list[dict[str, object]] = []

    class _FakeBackend:
        def extract(self, prepared) -> dict[str, object]:
            return {
                "service_date": "2025-06-25",
                "identity": "patient",
                "gender": "female",
                "ocr": {
                    "backend": "fake",
                    "raw_text": str(prepared.image_path),
                    "warnings": [],
                },
            }

    def fake_imwrite(path: str, frame: object) -> bool:
        Path(path).parent.mkdir(parents=True, exist_ok=True)
        Path(path).write_bytes(b"\x89PNG\r\n")
        return True

    def fake_resolve(
        explicit_dir=None,
        default_dir=None,
        *,
        env_overrides=None,
    ) -> _FakeBackend:
        calls.append(
            {
                "explicit_dir": explicit_dir,
                "default_dir": default_dir,
                "env_overrides": env_overrides,
            }
        )
        return _FakeBackend()

    monkeypatch.setenv("SCAN_DOC_PREPROCESS", "1")
    monkeypatch.setattr(
        "ocr_from2xlsx.app.filedialog.askdirectory", lambda **kwargs: str(output_dir)
    )
    monkeypatch.setenv("OCR_BACKEND", "plugin")
    monkeypatch.setattr(
        "ocr_from2xlsx.plugin_backend.PluginOcrBackend.resolve",
        fake_resolve,
    )
    monkeypatch.setitem(sys.modules, "cv2", SimpleNamespace(imwrite=fake_imwrite))

    ReviewApp._recognize_capture(app, frame=object())

    assert calls == [
        {
            "explicit_dir": None,
            "default_dir": None,
            "env_overrides": {"SCAN_DOC_PREPROCESS": "1"},
        }
    ]


def test_recognize_capture_uses_unique_names_when_output_dir_has_existing_scan_files(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    output_dir = tmp_path / "scan-output"
    output_dir.mkdir(parents=True, exist_ok=True)
    legacy_capture = output_dir / "scan-capture.png"
    legacy_capture.write_bytes(b"older capture")
    legacy_json = output_dir / "scan-prepared.json"
    legacy_json.write_text('{"legacy": true}', encoding="utf-8")

    class _FakeBackend:
        def extract(self, prepared) -> dict[str, object]:
            return {
                "service_date": "2025-06-25",
                "identity": "patient",
                "gender": "female",
                "ocr": {
                    "backend": "fake",
                    "raw_text": str(prepared.image_path),
                    "warnings": [],
                },
            }

    def fake_imwrite(path: str, frame: object) -> bool:
        Path(path).write_bytes(b"new capture")
        return True

    monkeypatch.setattr(
        "ocr_from2xlsx.app.filedialog.askdirectory", lambda **kwargs: str(output_dir)
    )
    monkeypatch.setenv("OCR_BACKEND", "plugin")
    monkeypatch.setattr(
        "ocr_from2xlsx.plugin_backend.PluginOcrBackend.resolve",
        lambda explicit_dir=None, default_dir=None: _FakeBackend(),
    )
    monkeypatch.setitem(sys.modules, "cv2", SimpleNamespace(imwrite=fake_imwrite))

    ReviewApp._recognize_capture(app, frame=object())

    assert legacy_capture.read_bytes() == b"older capture"
    assert legacy_json.read_text(encoding="utf-8") == '{"legacy": true}'
    assert (output_dir / "scan-capture-2.png").read_bytes() == b"new capture"
    assert app.loaded_json_path == output_dir / "scan-prepared-2.json"
    assert app.records[0].source.image_path == "scan-capture-2.png"
    assert app.records[0].source.preprocessed_image_path == "scan-capture-2.png"
    assert (output_dir / "scan-prepared-2.json").is_file()


def test_confirm_current_blocked_does_not_advance_or_persist_unconfirmed_name(
    app: ReviewApp, tmp_path: Path
) -> None:
    record = make_record("scan-0001")
    next_record = make_record("scan-0002")
    record.ocr.warnings = ["name.unconfirmed"]
    record.ocr.raw_text = "癌症資源中心服務紀錄表\n姓名 王小明\n病歷號 6250712919"
    app.records = [record, next_record]
    app.current_index = 0
    app.correction_store_path = tmp_path / "name_corrections.jsonl"
    session = StubSession(
        AcceptResult(
            record_id=record.record_id,
            status="blocked",
            row_number=None,
            blockers=["name.unconfirmed"],
            warnings=[],
        )
    )
    app.session = session
    app._show_record(record)

    ReviewApp._confirm_current(app)

    assert app.current_index == 0
    assert app.fields["record_id"].get() == record.record_id
    assert app.written_indices == set()
    assert record.ocr.warnings == ["name.unconfirmed"]
    assert session.calls == [(record.record_id, False, True)]
    assert load_corrections(app.correction_store_path) == []


def test_confirm_current_blocked_warns_user_with_blockers(
    app: ReviewApp, tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    record = make_record("scan-0001")
    app.records = [record]
    app.current_index = 0
    app.correction_store_path = tmp_path / "name_corrections.jsonl"
    app.session = StubSession(
        AcceptResult(
            record_id=record.record_id,
            status="blocked",
            row_number=None,
            blockers=["service_date.invalid", "patient.source.required"],
            warnings=[],
        )
    )
    app._show_record(record)
    shown: list = []
    monkeypatch.setattr("ocr_from2xlsx.app.messagebox.showwarning", lambda *a, **k: shown.append(a))

    ReviewApp._confirm_current(app)

    assert app.written_indices == set()
    assert shown, "a blocked confirm must warn the user it was not written"
    assert any("service_date.invalid" in str(a) for a in shown)


def test_confirm_blocks_empty_unconfirmed_name(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    record = make_record("scan-0001")
    record.ocr.warnings = ["name.unconfirmed"]
    app.records = [record]
    app.current_index = 0
    app.session = StubSession(
        AcceptResult(record_id=record.record_id, status="written", row_number=2, blockers=[], warnings=[])
    )
    app._show_record(record)
    app.fields["name"].set("")  # operator left the待確認 name blank
    warned: list = []
    monkeypatch.setattr("ocr_from2xlsx.app.messagebox.showwarning", lambda *a, **k: warned.append(a))

    ReviewApp._confirm_current(app)

    # Empty + unconfirmed must NOT be written or silently marked confirmed.
    assert app.session.calls == []
    assert app.written_indices == set()
    assert record.ocr.warnings == ["name.unconfirmed"]
    assert warned


def test_confirm_current_no_session_surfaces_error(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    # The button is no longer greyed out with no workbook loaded; pressing it must SURFACE the
    # reason instead of doing nothing (#confirm-required-fields).
    app.session = None
    app.records = [make_record("scan-0001")]
    app.current_index = 0
    errors: list = []
    monkeypatch.setattr("ocr_from2xlsx.app.messagebox.showerror", lambda *a, **k: errors.append(a))

    ReviewApp._confirm_current(app)

    assert errors and "缺少工作檔" in str(errors[0])


def test_confirm_current_blocks_empty_name_even_when_confirmed(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    # The name guard now applies to EVERY record, not only name.unconfirmed ones.
    record = make_record("scan-0001")
    app.records = [record]
    app.current_index = 0
    app.session = StubSession(
        AcceptResult(record_id=record.record_id, status="written", row_number=2, blockers=[], warnings=[])
    )
    app._show_record(record)
    app.fields["name"].set("   ")  # operator cleared the name
    warned: list = []
    monkeypatch.setattr("ocr_from2xlsx.app.messagebox.showwarning", lambda *a, **k: warned.append(a))

    ReviewApp._confirm_current(app)

    assert app.session.calls == []
    assert app.written_indices == set()
    assert warned


def test_confirm_current_writes_despite_blank_optional_fields(
    app: ReviewApp, tmp_path: Path
) -> None:
    # End-to-end: a name-only record (blank date / identity / gender) still writes (relaxed).
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    session = ImportSession.start(template, working)
    try:
        record = make_record("scan-0001")
        record.service_date = ""
        record.identity = ""
        record.gender = ""
        app.records = [record, make_record("scan-0002")]
        app.current_index = 0
        app.session = session
        app._show_record(record)
        app.fields["name"].set("王小明")

        ReviewApp._confirm_current(app)

        assert app.written_indices == {0}
        wb = load_workbook(working)
        try:
            sheet = wb["個案總表"]
            name_col = _column_for_header(sheet, BASIC_COLUMN_BY_FIELD["name"])
            assert sheet.cell(row=2, column=name_col).value == "王小明"
        finally:
            wb.close()
    finally:
        session.close()


def test_force_write_failure_does_not_persist_unconfirmed_name(app: ReviewApp, tmp_path: Path) -> None:
    record = make_record("scan-0001")
    record.ocr.warnings = ["name.unconfirmed"]
    record.ocr.raw_text = "王小明"
    app.records = [record]
    app.current_index = 0
    app.correction_store_path = tmp_path / "name_corrections.jsonl"
    app._show_record(record)

    class FailingSession:
        def accept_scan(
            self, record, force: bool = False, human_confirmed: bool = False,
            overwrite_row: int | None = None, relaxed: bool = False,
        ) -> AcceptResult:
            raise OSError("disk full")

        def close(self) -> None:
            return None

    app.session = FailingSession()

    app._force_write()

    assert record.ocr.warnings == ["name.unconfirmed"]
    assert load_corrections(app.correction_store_path) == []


def test_confirm_current_successful_write_tolerates_correction_store_failure(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    records = [make_record("scan-0001"), make_record("scan-0002")]
    records[0].ocr.warnings = ["name.unconfirmed"]
    app.records = records
    app.current_index = 0
    app.correction_store_path = Path("C:\\fake\\name_corrections.jsonl")
    app._show_record(records[0])
    app.session = StubSession(
        AcceptResult(
            record_id=records[0].record_id,
            status="written",
            row_number=2,
            blockers=[],
            warnings=[],
        )
    )
    errors: list[tuple[str, str]] = []
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showerror",
        lambda title, message: errors.append((title, message)),
    )

    def _raise_store_locked(**kwargs):
        raise OSError("store locked")

    monkeypatch.setattr("ocr_from2xlsx.app.confirm_name", _raise_store_locked)

    ReviewApp._confirm_current(app)

    assert errors == [("寫入失敗", "store locked")]
    assert app.written_indices == {0}
    assert app.current_index == 1
    assert app.fields["record_id"].get() == records[1].record_id


def test_confirm_and_force_write_tolerate_end_of_record_list(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    record = make_record("scan-0001")
    app.records = [record]
    app.current_index = 0
    app._show_record(record)
    app.editing = True
    app.session = StubSession(
        AcceptResult(
            record_id=record.record_id,
            status="written",
            row_number=2,
            blockers=[],
            warnings=[],
        )
    )
    infos: list[tuple[str, str]] = []
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showinfo",
        lambda title, message: infos.append((title, message)),
    )

    ReviewApp._confirm_current(app)
    ReviewApp._confirm_current(app)
    ReviewApp._force_write(app)

    assert app.current_index == 1
    assert app.session.calls == [(record.record_id, False, True)]
    assert infos


def test_previous_record_recovers_from_end_of_record_sentinel(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch
) -> None:
    record = make_record("scan-0001")
    app.records = [record]
    app.current_index = 0
    app._show_record(record)
    app.session = StubSession(
        AcceptResult(
            record_id=record.record_id,
            status="written",
            row_number=2,
            blockers=[],
            warnings=[],
        )
    )
    infos: list[tuple[str, str]] = []
    monkeypatch.setattr(
        "ocr_from2xlsx.app.messagebox.showinfo",
        lambda title, message: infos.append((title, message)),
    )

    ReviewApp._confirm_current(app)
    ReviewApp._next_record(app)
    ReviewApp._previous_record(app)

    assert app.current_index == 0
    assert app.fields["record_id"].get() == record.record_id
    assert infos


def test_choose_template_single_dialog_defaults_output_and_clears_indices(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # 開啟報表 is now ONE selection (the source report). The working file defaults under the
    # output root — no second "輸出資料夾" prompt (#single-folder-prompt).
    app.written_indices = {0, 1}
    app.output_root = tmp_path  # default 'output' overridden for test isolation
    closed: list[bool] = []

    class ExistingSession:
        def close(self) -> None:
            closed.append(True)

    app.session = ExistingSession()
    template_path = str(tmp_path / "base.xlsx")
    start_calls: list[tuple[str, Path]] = []
    askdir_calls: list[object] = []

    monkeypatch.setattr(
        "ocr_from2xlsx.app.filedialog.askopenfilename", lambda **kwargs: template_path
    )
    monkeypatch.setattr(
        "ocr_from2xlsx.app.filedialog.askdirectory",
        lambda **kwargs: askdir_calls.append(kwargs) or "",
    )

    class NewSession:
        def close(self) -> None:
            return None

    def fake_start(template: str, working: Path) -> NewSession:
        start_calls.append((template, working))
        return NewSession()

    monkeypatch.setattr("ocr_from2xlsx.app.ImportSession.start", fake_start)

    app._choose_template()

    assert closed == [True]
    assert askdir_calls == []  # no second folder prompt
    assert start_calls == [(template_path, tmp_path / "匯入中.xlsx")]
    assert app.written_indices == set()


def test_import_folder_batch_single_dialog_defaults_output(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # 匯入資料夾 is now ONE selection (the photo/PDF source). Output defaults under the output
    # root — no second "輸出資料夾" prompt (#single-folder-prompt).
    app.output_root = tmp_path
    askdir_calls: list[object] = []
    captured: dict[str, object] = {}

    monkeypatch.setattr(
        "ocr_from2xlsx.app.filedialog.askdirectory",
        lambda **kwargs: (askdir_calls.append(kwargs) or str(tmp_path / "photos")),
    )
    monkeypatch.setattr(
        "ocr_from2xlsx.scan.next_output_artifact_path",
        lambda out_dir, name: (captured.__setitem__("out", out_dir) or (Path(out_dir) / name)),
    )
    monkeypatch.setattr(app, "_run_recognition_async", lambda *a, **k: captured.__setitem__("ran", True))

    app._import_folder_batch()

    assert len(askdir_calls) == 1  # only the source-folder prompt, not in+out
    assert Path(captured["out"]) == tmp_path
    assert captured.get("ran") is True


def test_resolve_output_dir_override_dev_and_frozen(
    monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # #single-folder-prompt: output is fixed + predictable (exe dir when frozen, cwd in dev),
    # never a surprise cwd. output_root overrides for tests.
    app = ReviewApp.__new__(ReviewApp)

    app.output_root = tmp_path / "custom"
    assert app._resolve_output_dir() == tmp_path / "custom"

    app.output_root = None
    monkeypatch.chdir(tmp_path)
    monkeypatch.setattr(app_module.sys, "frozen", False, raising=False)
    assert app._resolve_output_dir() == Path.cwd() / "output"

    monkeypatch.setattr(app_module.sys, "frozen", True, raising=False)
    monkeypatch.setattr(app_module.sys, "executable", str(tmp_path / "bin" / "ocr.exe"), raising=False)
    frozen = app._resolve_output_dir()
    assert frozen.name == "output" and frozen.parent.name == "bin"


def test_end_to_end_roundtrip_self_check() -> None:
    """Run the auditable GUI round-trip self-check (build/verify_roundtrip.py) as a suite gate:
    golden JSON -> ConfirmForm load/write -> XLSX -> assert it matches. main() returns 0 on
    success (and SKIPs to 0 when no Tk display is available, e.g. headless CI)."""
    import importlib.util

    script = Path(__file__).resolve().parents[1] / "build" / "verify_roundtrip.py"
    spec = importlib.util.spec_from_file_location("verify_roundtrip", script)
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    assert module.main() == 0


def test_add_blank_record_appends_and_shows_for_manual_entry(app: ReviewApp) -> None:
    # #manual-blank-record: 新增頁面 builds a blank record on demand (no JSON/scan needed) and
    # shows it for filling. Auto-numbered manual-NNNN ids; works even before 開啟報表.
    app.records = []
    app.current_index = -1
    app.session = None

    ReviewApp._add_blank_record(app)
    assert len(app.records) == 1
    rec = app.records[0]
    assert rec.record_id == "manual-0001"
    assert (rec.name, rec.identity, rec.service_date, rec.gender) == ("", "", "", "")
    assert app.current_index == 0
    assert app.fields["record_id"].get() == "manual-0001"

    ReviewApp._add_blank_record(app)
    assert app.records[1].record_id == "manual-0002"
    assert app.current_index == 1


def test_add_blank_record_then_confirm_writes_manually(app: ReviewApp, tmp_path: Path) -> None:
    # End-to-end manual entry: 開啟報表 (session) -> 新增頁面 -> fill name -> 確認並寫入 -> XLSX,
    # with no JSON load anywhere.
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    session = ImportSession.start(template, working)
    try:
        app.session = session
        app.records = []
        app.current_index = -1

        ReviewApp._add_blank_record(app)
        app.fields["name"].set("王小明")
        ReviewApp._confirm_current(app)

        assert app.written_indices == {0}
        wb = load_workbook(working)
        try:
            sheet = wb["個案總表"]
            name_col = _column_for_header(sheet, BASIC_COLUMN_BY_FIELD["name"])
            assert sheet.cell(row=2, column=name_col).value == "王小明"
        finally:
            wb.close()
    finally:
        session.close()


def test_choose_template_auto_creates_blank_record_for_manual_entry(
    app: ReviewApp, monkeypatch: pytest.MonkeyPatch, tmp_path: Path
) -> None:
    # #manual-blank-record: 開啟報表 with no records loaded starts the operator on a blank record,
    # so they can fill and write immediately — no extra click, no JSON/scan.
    app.output_root = tmp_path
    app.records = []
    app.current_index = -1
    monkeypatch.setattr(
        "ocr_from2xlsx.app.filedialog.askopenfilename", lambda **k: str(tmp_path / "t.xlsx")
    )

    class _Sess:
        def close(self) -> None:
            return None

    monkeypatch.setattr("ocr_from2xlsx.app.ImportSession.start", lambda *a, **k: _Sess())

    app._choose_template()

    assert len(app.records) == 1
    assert app.records[0].record_id == "manual-0001"
    assert app.current_index == 0
    assert app.fields["record_id"].get() == "manual-0001"


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("2026/06/16", "2026-06-16"),  # slash Gregorian — what operators actually type
        ("2026-06-16", "2026-06-16"),  # already ISO — unchanged
        ("2026.6.16", "2026-06-16"),   # dotted, single-digit month/day
        ("115/6/28", "2026-06-28"),    # ROC (民國) year < 1911 → +1911
        ("", ""),                       # empty stays empty (no date → no month)
        ("not a date", "not a date"),  # unparseable kept as-is (don't lose input)
    ],
)
def test_normalize_service_date_handles_slash_and_roc(raw: str, expected: str) -> None:
    # #manual-date-month: the manual form date field is free text titled 服務年/月/日; operators type
    # slash Gregorian (2026/06/16) or ROC (115/6/28). Normalize to ISO so service_month_label works.
    from ocr_from2xlsx.domain import normalize_service_date

    assert normalize_service_date(raw) == expected


def test_manual_slash_date_writes_service_month(app: ReviewApp, tmp_path: Path) -> None:
    # #manual-date-month regression: 開啟報表 -> 填 2026/06/16 -> 確認並寫入 must write 服務月份=6月
    # (it was blank). Real template + read-back: service_month_label needs ISO, so the manual path
    # must normalize the typed slash date (the scan path already does via parse_roc_date).
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    session = ImportSession.start(template, working)
    try:
        app.session = session
        app.records = []
        app.current_index = -1
        app.loaded_json_path = None

        ReviewApp._add_blank_record(app)
        app.fields["name"].set("王小明")
        app.fields["service_date"].set("2026/06/16")
        ReviewApp._confirm_current(app)

        wb = load_workbook(working)
        try:
            sheet = wb["個案總表"]
            month_col = _column_for_header(sheet, BASIC_COLUMN_BY_FIELD["service_month"])
            date_col = _column_for_header(sheet, BASIC_COLUMN_BY_FIELD["service_date"])
            assert sheet.cell(row=2, column=month_col).value == "6月"
            assert sheet.cell(row=2, column=date_col).value == "2026-06-16"
        finally:
            wb.close()
    finally:
        session.close()


def test_confirm_manual_mode_presents_fresh_blank_for_next_sheet(
    app: ReviewApp, tmp_path: Path
) -> None:
    # #manual-continue regression: after 開啟報表 -> 填單 -> 確認並寫入, manual mode must leave the
    # operator on a fresh blank page so the NEXT sheet just works. Previously _next_record ran the
    # index out of bounds and the following 確認並寫入 wrongly errored 「請先載入 JSON 資料。」.
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    session = ImportSession.start(template, working)
    try:
        app.session = session
        app.records = []
        app.current_index = -1
        app.loaded_json_path = None  # manual mode (no JSON/scan)

        ReviewApp._add_blank_record(app)  # the auto blank from 開啟報表
        app.fields["name"].set("王小明")
        ReviewApp._confirm_current(app)

        assert app.written_indices == {0}
        # A fresh blank page is active and in-bounds → next 確認並寫入 won't hit the JSON guard.
        assert len(app.records) == 2
        assert app.current_index == 1
        assert 0 <= app.current_index < len(app.records)
        assert app.records[1].name == ""
        assert app.records[1].record_id.startswith("manual-")
    finally:
        session.close()


def test_confirm_json_mode_does_not_auto_add_blank_at_end(
    app: ReviewApp, tmp_path: Path
) -> None:
    # Guard for the Bug 1 fix: auto-blank-on-confirm is manual-mode ONLY. With a JSON/scan source
    # loaded, confirming the LAST record must NOT append a stray blank — it stays "no more records".
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    session = ImportSession.start(template, working)
    try:
        app.session = session
        app.records = [make_record("r1")]
        app.current_index = 0
        app.loaded_json_path = tmp_path / "batch.json"  # JSON/scan mode
        app.fields["name"].set("王小明")
        ReviewApp._confirm_current(app)

        assert app.written_indices == {0}
        assert len(app.records) == 1  # no stray blank appended
        assert app.current_index >= len(app.records)  # past end = "no more records"
    finally:
        session.close()


def test_blank_page_focuses_service_date_not_name() -> None:
    # #focus-service-date: a fresh blank page (開啟報表 / after 確認並寫入) must put the caret on
    # 服務日期 (the top field) for natural top-down entry — not jump down to 姓名. Uses a real
    # ReviewApp because the headless FakeConfirmForm has no _focus to exercise.
    try:
        app = ReviewApp()
    except tk.TclError as exc:
        pytest.skip(f"no display available for Tk: {exc}")
    app.withdraw()
    try:
        app.records = []
        app.current_index = -1
        app._add_blank_record()
        assert app.confirm_form._current_focus == "service_date"
    finally:
        app.destroy()


def test_window_starts_maximized() -> None:
    # #startup-maximized: the app should open maximized so the wide 個案總表 form/preview use the
    # whole screen, instead of the small 1200x720 default.
    try:
        app = ReviewApp()
    except tk.TclError as exc:
        pytest.skip(f"no display available for Tk: {exc}")
    try:
        app.update_idletasks()
        assert app.state() == "zoomed"
    finally:
        app.destroy()
