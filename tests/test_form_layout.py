from __future__ import annotations

from ocr_from2xlsx.form_layout import Field, FormLayout, Option, Section, service_record_layout


def _tiny_layout() -> FormLayout:
    """Minimal layout for testing accessors."""
    return FormLayout(
        template_id="t",
        sections=(
            Section(
                id="B",
                title="綜合身份統計",
                fields=(
                    Field(
                        key="gender",
                        title="性別",
                        kind="single_choice",
                        record_path="gender",
                        anchor_cell="A25",
                        options=(
                            Option(label="女性", code="female", cell="B25"),
                            Option(label="男性", code="male", cell="B26"),
                        ),
                    ),
                    Field(
                        key="name",
                        title="姓名",
                        kind="text",
                        record_path="name",
                        anchor_cell="B23",
                        options=(),
                    ),
                    Field(
                        key="page_num",
                        title="頁碼",
                        kind="text",
                        record_path=None,
                        anchor_cell="Z1",
                        options=(),
                    ),
                ),
            ),
        ),
    )


def test_field_by_key_and_iter() -> None:
    layout = _tiny_layout()
    
    # field_by_key lookups
    assert layout.field_by_key("gender").kind == "single_choice"
    assert layout.field_by_key("name").options == ()
    assert layout.field_by_key("missing") is None
    
    # iter_fields yields Field objects
    field_keys = [f.key for f in layout.iter_fields()]
    assert field_keys == ["gender", "name", "page_num"]


def test_field_with_no_record_counterpart() -> None:
    layout = _tiny_layout()
    
    # Fields can have record_path=None when no Record counterpart exists
    page_num = layout.field_by_key("page_num")
    assert page_num is not None
    assert page_num.record_path is None


def test_iter_options_and_options_by_code() -> None:
    layout = _tiny_layout()
    
    # iter_options returns (Field, Option) pairs
    option_pairs = [(f.key, o.code) for f, o in layout.iter_options()]
    assert option_pairs == [("gender", "female"), ("gender", "male")]
    
    # options_by_code returns dict[code, Option] for a field
    gender_opts = layout.options_by_code("gender")
    assert gender_opts["female"].cell == "B25"
    
    # options_by_code for text field returns empty dict
    name_opts = layout.options_by_code("name")
    assert name_opts == {}


def test_options_by_code_fails_on_missing_field() -> None:
    layout = _tiny_layout()
    
    # options_by_code should raise KeyError for unknown field keys
    try:
        layout.options_by_code("missing")
        assert False, "Expected KeyError but call succeeded"
    except KeyError as e:
        assert "missing" in str(e)


def test_dataclass_sequences_are_coerced_to_tuples() -> None:
    """Verify that passing lists to options/fields/sections doesn't allow mutation."""
    # Build layout with Python lists
    options_list = [
        Option(label="女性", code="female", cell="B25"),
        Option(label="男性", code="male", cell="B26"),
    ]
    fields_list = [
        Field(
            key="gender",
            title="性別",
            kind="single_choice",
            record_path="gender",
            anchor_cell="A25",
            options=options_list,
        ),
    ]
    sections_list = [
        Section(id="B", title="綜合身份統計", fields=fields_list),
    ]
    layout = FormLayout(template_id="t", sections=sections_list)
    
    # Stored attributes must be tuples, not lists
    assert isinstance(layout.sections, tuple)
    assert isinstance(layout.sections[0].fields, tuple)
    assert isinstance(layout.sections[0].fields[0].options, tuple)
    
    # Mutating original lists should not affect stored model
    original_gender_field = layout.sections[0].fields[0]
    original_options_count = len(original_gender_field.options)
    
    options_list.append(Option(label="其他", code="other", cell="B27"))
    fields_list.append(
        Field(key="name", title="姓名", kind="text", record_path="name", anchor_cell="B23")
    )
    sections_list.append(Section(id="C", title="另一節", fields=()))
    
    # Stored model should not change
    assert len(layout.sections) == 1
    assert len(layout.sections[0].fields) == 1
    assert len(layout.sections[0].fields[0].options) == original_options_count


def test_duplicate_option_code_raises_error() -> None:
    """Verify that duplicate Option.code values within a field raise ValueError."""
    try:
        Field(
            key="status",
            title="狀態",
            kind="single_choice",
            record_path="status",
            anchor_cell="A1",
            options=(
                Option(label="Active", code="active", cell="B1"),
                Option(label="Active (other)", code="active", cell="B2"),
            ),
        )
        assert False, "Expected ValueError for duplicate option code"
    except ValueError as e:
        assert "active" in str(e).lower() or "duplicate" in str(e).lower()


def test_duplicate_field_key_raises_error() -> None:
    """Verify that duplicate Field.key values across the layout raise ValueError."""
    try:
        FormLayout(
            template_id="t",
            sections=(
                Section(
                    id="A",
                    title="Section A",
                    fields=(
                        Field(key="id", title="ID", kind="text", record_path="id", anchor_cell="A1"),
                    ),
                ),
                Section(
                    id="B",
                    title="Section B",
                    fields=(
                        Field(key="id", title="ID (copy)", kind="text", record_path="id2", anchor_cell="A2"),
                    ),
                ),
            ),
        )
        assert False, "Expected ValueError for duplicate field key"
    except ValueError as e:
        assert "id" in str(e).lower() or "duplicate" in str(e).lower()


def test_constructor_annotations_accept_sequences() -> None:
    """Verify that constructor signatures advertise support for list/sequence inputs.
    
    This is a regression test for the API contract: the implementation accepts
    lists/sequences and coerces them to tuples internally, so the public
    constructor signature should reflect that callers can pass lists.
    """
    import inspect
    from collections.abc import Sequence
    
    # Field(options=...) should accept Sequence, not just tuple
    field_sig = inspect.signature(Field)
    options_param = field_sig.parameters["options"]
    options_annotation = options_param.annotation
    
    # The annotation should allow Sequence[Option] (or similar), not just tuple[Option, ...]
    # We check that it's not strictly tuple by looking at the annotation string representation
    # or by checking if it's a Sequence type hint
    assert options_annotation != "tuple[Option, ...]", (
        f"Field.options annotation should accept sequences, not tuple-only: {options_annotation}"
    )
    
    # Section(fields=...) should accept Sequence, not just tuple
    section_sig = inspect.signature(Section)
    fields_param = section_sig.parameters["fields"]
    fields_annotation = fields_param.annotation
    
    assert fields_annotation != "tuple[Field, ...]", (
        f"Section.fields annotation should accept sequences, not tuple-only: {fields_annotation}"
    )
    
    # FormLayout(sections=...) should accept Sequence, not just tuple
    layout_sig = inspect.signature(FormLayout)
    sections_param = layout_sig.parameters["sections"]
    sections_annotation = sections_param.annotation
    
    assert sections_annotation != "tuple[Section, ...]", (
        f"FormLayout.sections annotation should accept sequences, not tuple-only: {sections_annotation}"
    )


def test_field_options_has_default_in_constructor_and_dataclass_metadata() -> None:
    """Verify that Field.options has a default both in constructor and dataclass metadata.
    
    Regression test: Field.__init__ has options=() default, so dataclass metadata
    should also report a default (not MISSING). Both normal construction and
    dataclass introspection must agree on the public contract.
    """
    import dataclasses
    import inspect
    
    # Constructor signature should show a default
    field_sig = inspect.signature(Field)
    options_param = field_sig.parameters["options"]
    assert options_param.default is not inspect.Parameter.empty, (
        "Field.__init__(options=...) should have a default value in signature"
    )
    
    # Dataclass metadata should also show a default (not MISSING)
    field_info = next(f for f in dataclasses.fields(Field) if f.name == "options")
    has_default = (
        field_info.default is not dataclasses.MISSING
        or field_info.default_factory is not dataclasses.MISSING
    )
    assert has_default, (
        "Field.options dataclass field should have a default or default_factory"
    )


def test_text_field_with_options_raises_error() -> None:
    """Verify that text fields with non-empty options raise ValueError."""
    try:
        Field(
            key="name",
            title="姓名",
            kind="text",
            record_path="name",
            anchor_cell="B23",
            options=(Option(label="N/A", code="na", cell="B24"),),
        )
        assert False, "Expected ValueError for text field with options"
    except ValueError as e:
        assert "text" in str(e).lower() or "options" in str(e).lower()


def test_single_choice_field_without_options_raises_error() -> None:
    """Verify that single_choice fields without options raise ValueError."""
    try:
        Field(
            key="status",
            title="狀態",
            kind="single_choice",
            record_path="status",
            anchor_cell="A1",
            options=(),
        )
        assert False, "Expected ValueError for single_choice field without options"
    except ValueError as e:
        assert "single_choice" in str(e).lower() or "options" in str(e).lower()


def test_multi_choice_field_without_options_raises_error() -> None:
    """Verify that multi_choice fields without options raise ValueError."""
    try:
        Field(
            key="interests",
            title="興趣",
            kind="multi_choice",
            record_path="interests",
            anchor_cell="A10",
            options=(),
        )
        assert False, "Expected ValueError for multi_choice field without options"
    except ValueError as e:
        assert "multi_choice" in str(e).lower() or "options" in str(e).lower()


def test_field_with_invalid_kind_raises_error() -> None:
    """Verify that Field rejects invalid kind values at runtime."""
    try:
        Field(
            key="test",
            title="測試",
            kind="bogus",
            record_path="test",
            anchor_cell="A1",
            options=(),
        )
        assert False, "Expected ValueError for invalid kind='bogus'"
    except ValueError as e:
        assert "kind" in str(e).lower() or "bogus" in str(e).lower()


# --- Task 2: service_record_layout() tests ---


def test_layout_covers_expected_fields():
    layout = service_record_layout()
    keys = {f.key for f in layout.iter_fields()}
    assert {
        "service_date", "identity", "name", "medical_record_no", "gender",
        "nationality", "age", "channel", "disease_status", "source", "cancer",
        "newly_diagnosed", "supplies", "internal_referrals", "external_referrals",
        "referral_outcomes",
    } <= keys
    assert "consultation.health_medical" in keys


def test_choice_option_counts_and_record_paths():
    layout = service_record_layout()
    assert len(layout.field_by_key("cancer").options) == 25
    assert len(layout.field_by_key("age").options) == 7
    assert layout.field_by_key("age").record_path == "patient_fields.age_group"
    assert layout.field_by_key("identity").record_path == "identity"
    assert layout.field_by_key("cancer").record_path == "patient_fields.cancers"
    assert layout.field_by_key("consultation.health_medical").record_path == "services.consultation.health_medical"
    assert layout.field_by_key("diagnosis_date").record_path is None


def test_option_codes_are_constants_legal():
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    assert {o.code for f, o in layout.iter_options() if f.key == "identity"} <= constants.IDENTITIES
    assert {o.code for f, o in layout.iter_options() if f.key == "gender"} <= constants.GENDERS
    assert set(layout.options_by_code("cancer")) <= set(constants.CANCER_LABELS)


# --- Issue 1: Generic record value normalization contract ---


def test_field_selected_codes_for_boolean_field():
    """Test generic value normalization for boolean fields like newly_diagnosed."""
    layout = service_record_layout()
    fld = layout.field_by_key("newly_diagnosed")
    
    # True maps to ("true",)
    assert fld.selected_codes(True) == ("true",)
    
    # False and None map to empty tuple
    assert fld.selected_codes(False) == ()
    assert fld.selected_codes(None) == ()


def test_field_selected_codes_for_string_single_choice():
    """Test generic value normalization for string-coded single-choice fields."""
    layout = service_record_layout()
    gender = layout.field_by_key("gender")
    
    # String value maps to tuple containing that code
    assert gender.selected_codes("female") == ("female",)
    assert gender.selected_codes("male") == ("male",)
    
    # None maps to empty tuple
    assert gender.selected_codes(None) == ()


def test_field_selected_codes_for_string_multi_choice():
    """Test generic value normalization for string-coded multi-choice fields."""
    layout = service_record_layout()
    cancer = layout.field_by_key("cancer")
    
    # List of strings maps to tuple of codes
    assert cancer.selected_codes(["brain_cancer", "lung_cancer"]) == ("brain_cancer", "lung_cancer")
    
    # Empty list or None maps to empty tuple
    assert cancer.selected_codes([]) == ()
    assert cancer.selected_codes(None) == ()


def test_field_selected_codes_for_text_field():
    """Text fields should not support selected_codes (or return empty tuple)."""
    layout = service_record_layout()
    name = layout.field_by_key("name")
    
    # Text fields don't have options, should return empty tuple or raise
    assert name.selected_codes("John Doe") == ()


def test_field_selected_codes_bool_only_for_bool_backed_single_choice():
    """Bool values should only work for bool-backed single-choice fields (codes subset of {true, false})."""
    import pytest
    layout = service_record_layout()
    
    # newly_diagnosed is bool-backed (only has "true" option)
    newly_diagnosed = layout.field_by_key("newly_diagnosed")
    assert newly_diagnosed.selected_codes(True) == ("true",)
    assert newly_diagnosed.selected_codes(False) == ()
    assert newly_diagnosed.selected_codes(None) == ()
    
    # gender is NOT bool-backed (has "female", "male", "other")
    gender = layout.field_by_key("gender")
    with pytest.raises(TypeError, match=r"bool.*not.*bool-backed"):
        gender.selected_codes(True)
    
    # cancer is multi_choice, should also reject bool
    cancer = layout.field_by_key("cancer")
    with pytest.raises(TypeError, match=r"bool.*not.*bool-backed"):
        cancer.selected_codes(True)


def test_field_selected_codes_list_only_for_multi_choice():
    """List values should only work for multi_choice fields, not single_choice."""
    import pytest
    layout = service_record_layout()
    
    # cancer is multi_choice, accepts lists
    cancer = layout.field_by_key("cancer")
    assert cancer.selected_codes(["brain_cancer"]) == ("brain_cancer",)
    
    # gender is single_choice, should reject lists
    gender = layout.field_by_key("gender")
    with pytest.raises(TypeError, match=r"list.*single_choice"):
        gender.selected_codes(["female", "male"])


def test_field_selected_codes_bool_backed_maps_by_code_not_order():
    """Bool-backed field with both true and false codes must map by code value, not option order."""
    import pytest
    from ocr_from2xlsx.form_layout import Field, Option
    
    # Create a bool-backed field with options in reverse order: false, then true
    field_reverse_order = Field(
        key="reversed_bool",
        title="Reversed Boolean Field",
        kind="single_choice",
        record_path="reversed_bool",
        anchor_cell="A1",
        options=(
            Option(label="No", code="false", cell="A1"),
            Option(label="Yes", code="true", cell="A2"),
        ),
    )
    
    # True must map to "true" code, not the first option
    assert field_reverse_order.selected_codes(True) == ("true",)
    # False must map to "false" code, not empty tuple
    assert field_reverse_order.selected_codes(False) == ("false",)
    # None still maps to empty tuple
    assert field_reverse_order.selected_codes(None) == ()


def test_field_selected_codes_newly_diagnosed_preserves_behavior():
    """Verify newly_diagnosed (only has 'true' option) still behaves correctly."""
    layout = service_record_layout()
    fld = layout.field_by_key("newly_diagnosed")
    
    # True maps to ("true",)
    assert fld.selected_codes(True) == ("true",)
    # False maps to () because there's no "false" option
    assert fld.selected_codes(False) == ()
    # None maps to ()
    assert fld.selected_codes(None) == ()


def test_field_selected_codes_rejects_invalid_scalar_types():
    """Invalid scalar types like int should raise TypeError."""
    import pytest
    layout = service_record_layout()
    
    gender = layout.field_by_key("gender")
    with pytest.raises(TypeError, match=r"int.*not supported"):
        gender.selected_codes(123)
    
    cancer = layout.field_by_key("cancer")
    with pytest.raises(TypeError, match=r"int.*not supported"):
        cancer.selected_codes(456)


def test_field_selected_codes_rejects_invalid_list_members():
    """Lists containing non-string elements should raise TypeError."""
    import pytest
    layout = service_record_layout()
    
    cancer = layout.field_by_key("cancer")
    
    # List with bool should fail
    with pytest.raises(TypeError, match=r"list.*bool"):
        cancer.selected_codes([True])
    
    # List with int should fail
    with pytest.raises(TypeError, match=r"list.*int"):
        cancer.selected_codes([1, "brain_cancer"])
    
    # Mixed valid and invalid should also fail
    with pytest.raises(TypeError, match=r"list.*int"):
        cancer.selected_codes(["lung_cancer", 42])


def test_check_option_codes_match_constants_verifies_exact_labels():
    """Verify _check_option_codes_match_constants checks exact label equality when expected_labels provided."""
    import pytest
    
    # Create a field with correct codes but WRONG label for one option
    wrong_label_field = Field(
        key="test_field",
        title="Test",
        kind="single_choice",
        record_path="test",
        anchor_cell="A1",
        options=(
            Option(label="正確標籤", code="code1", cell="A1"),
            Option(label="錯誤標籤", code="code2", cell="A2"),  # Wrong label!
        ),
    )
    
    expected_codes = {"code1", "code2"}
    expected_labels = {"code1": "正確標籤", "code2": "應該的標籤"}  # code2 should be "應該的標籤"
    
    # Should raise AssertionError because code2 has wrong label
    with pytest.raises(AssertionError, match=r"code2.*label"):
        _check_option_codes_match_constants(wrong_label_field, expected_codes, expected_labels)


# --- Issue 2: Exact section and field order, canonical code backing ---


def test_service_record_exact_section_order():
    """Section order must be: top, A, B, C."""
    layout = service_record_layout()
    section_ids = [sec.id for sec in layout.sections]
    assert section_ids == ["top", "A", "B", "C"]


def test_service_record_section_top_field_order():
    """Top section must contain exactly: service_date."""
    layout = service_record_layout()
    top = layout.sections[0]
    assert top.id == "top"
    field_keys = [f.key for f in top.fields]
    assert field_keys == ["service_date"]


def test_service_record_section_a_field_order():
    """Section A must contain consultation.* fields, then supplies, internal/external referrals, referral_outcomes."""
    layout = service_record_layout()
    section_a = layout.sections[1]
    assert section_a.id == "A"
    field_keys = [f.key for f in section_a.fields]
    assert field_keys == [
        "consultation.health_medical",
        "consultation.symptom_side_effect",
        "consultation.nutrition_diet",
        "consultation.psychosocial_emotion",
        "consultation.financial_social",
        "consultation.care_support",
        "supplies",
        "internal_referrals",
        "external_referrals",
        "referral_outcomes",
    ]


def test_service_record_section_b_field_order():
    """Section B must contain identity, name, medical_record_no, diagnosis_date, gender, nationality, age."""
    layout = service_record_layout()
    section_b = layout.sections[2]
    assert section_b.id == "B"
    field_keys = [f.key for f in section_b.fields]
    assert field_keys == [
        "identity",
        "name",
        "medical_record_no",
        "diagnosis_date",
        "gender",
        "nationality",
        "age",
    ]


def test_service_record_section_c_field_order():
    """Section C must contain channel, disease_status, source, cancer, newly_diagnosed."""
    layout = service_record_layout()
    section_c = layout.sections[3]
    assert section_c.id == "C"
    field_keys = [f.key for f in section_c.fields]
    assert field_keys == [
        "channel",
        "disease_status",
        "source",
        "cancer",
        "newly_diagnosed",
    ]


def _check_option_codes_match_constants(field, expected_codes, expected_labels=None):
    """Helper to verify option codes match canonical constants."""
    actual_codes = {opt.code for opt in field.options}
    assert actual_codes == expected_codes, (
        f"Field {field.key} option codes mismatch: expected {expected_codes}, got {actual_codes}"
    )
    # Also check that each code has corresponding label (if labels provided)
    if expected_labels:
        for opt in field.options:
            assert opt.code in expected_labels, f"Option code {opt.code!r} missing from labels"
            # Verify exact label equality
            assert opt.label == expected_labels[opt.code], (
                f"Field {field.key} option {opt.code!r}: "
                f"label mismatch: expected {expected_labels[opt.code]!r}, got {opt.label!r}"
            )


def test_identity_options_match_constants():
    """identity field options must match IDENTITIES and IDENTITY_LABELS."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("identity")
    _check_option_codes_match_constants(fld, constants.IDENTITIES, constants.IDENTITY_LABELS)


def test_gender_options_match_constants():
    """gender field options must match GENDERS and GENDER_LABELS."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("gender")
    _check_option_codes_match_constants(fld, constants.GENDERS, constants.GENDER_LABELS)


def test_nationality_options_match_constants():
    """nationality field options must match PATIENT_ENUMS['nationality'] and NATIONALITY_LABELS."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("nationality")
    _check_option_codes_match_constants(
        fld, constants.PATIENT_ENUMS["nationality"], constants.NATIONALITY_LABELS
    )


def test_age_options_match_constants():
    """age field options must match PATIENT_ENUMS['age_group'] and AGE_GROUP_LABELS."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("age")
    _check_option_codes_match_constants(
        fld, constants.PATIENT_ENUMS["age_group"], constants.AGE_GROUP_LABELS
    )


def test_channel_options_match_constants():
    """channel field options must match PATIENT_ENUMS['channel'] and CHANNEL_LABELS."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("channel")
    _check_option_codes_match_constants(
        fld, constants.PATIENT_ENUMS["channel"], constants.CHANNEL_LABELS
    )


def test_disease_status_options_match_constants():
    """disease_status field options must match PATIENT_ENUMS['disease_status'] and DISEASE_STATUS_LABELS."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("disease_status")
    _check_option_codes_match_constants(
        fld, constants.PATIENT_ENUMS["disease_status"], constants.DISEASE_STATUS_LABELS
    )


def test_source_options_match_constants():
    """source field options must match PATIENT_ENUMS['source'] and SOURCE_LABELS."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("source")
    _check_option_codes_match_constants(
        fld, constants.PATIENT_ENUMS["source"], constants.SOURCE_LABELS
    )


def test_cancer_options_match_constants():
    """cancer field options must match CANCER_LABELS."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("cancer")
    _check_option_codes_match_constants(fld, set(constants.CANCER_LABELS.keys()), constants.CANCER_LABELS)


def test_consultation_fields_match_service_categories():
    """All consultation.* fields must match SERVICE_CATEGORIES."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    
    consultation_mapping = {
        "consultation.health_medical": "health_medical",
        "consultation.symptom_side_effect": "symptom_side_effect",
        "consultation.nutrition_diet": "nutrition_diet",
        "consultation.psychosocial_emotion": "psychosocial_emotion",
        "consultation.financial_social": "financial_social",
        "consultation.care_support": "care_support",
    }
    
    for field_key, category_key in consultation_mapping.items():
        fld = layout.field_by_key(field_key)
        expected_codes = constants.SERVICE_CATEGORIES[category_key]
        actual_codes = {opt.code for opt in fld.options}
        assert actual_codes == expected_codes, (
            f"Field {field_key} codes mismatch: expected {expected_codes}, got {actual_codes}"
        )


def test_supplies_options_match_constants():
    """supplies field options must match SUPPLY_CODES."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("supplies")
    _check_option_codes_match_constants(fld, constants.SUPPLY_CODES)


def test_internal_referrals_options_match_constants():
    """internal_referrals field options must match RESOURCE_CODES."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("internal_referrals")
    _check_option_codes_match_constants(fld, constants.RESOURCE_CODES)


def test_external_referrals_options_match_constants():
    """external_referrals field options must match RESOURCE_CODES."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("external_referrals")
    _check_option_codes_match_constants(fld, constants.RESOURCE_CODES)


def test_referral_outcomes_options_match_constants():
    """referral_outcomes field options must match OUTCOME_CODES."""
    from ocr_from2xlsx import constants
    layout = service_record_layout()
    fld = layout.field_by_key("referral_outcomes")
    _check_option_codes_match_constants(fld, constants.OUTCOME_CODES)


# --- Task 3: Openpyxl-backed validation against the real workbook ---


# Workbook helpers
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

_XLSX = Path(__file__).resolve().parents[1] / "115年整年月報表統計_單案統計加總版(下拉式)(空白).xlsx"
_SHEET = "服務紀錄表"

# Aggregate count cells that contain checkboxes but are NOT option cells
_NON_OPTION_CHECKBOX_CELLS = {
    "C25",  # 女性數量
    "D25",  # 女性數量
    "C26",  # 男性數量
    "D26",  # 男性數量
    "C27",  # 其他數量
    "D27",  # 其他數量
}


def _read_workbook_cells():
    """Read all non-empty cell texts from the workbook sheet."""
    if load_workbook is None:
        import pytest
        pytest.skip("openpyxl not available")
    
    if not _XLSX.exists():
        import pytest
        pytest.skip(f"Workbook not found: {_XLSX}")
    
    wb = load_workbook(_XLSX, data_only=True)
    ws = wb[_SHEET]
    
    cells = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                text = str(cell.value).strip()
                if text:
                    cells[cell.coordinate] = text
    
    wb.close()
    return cells


def test_every_modeled_option_matches_sheet_cell():
    """Every modeled option must correspond to a real checkbox cell in the workbook."""
    layout = service_record_layout()
    cells = _read_workbook_cells()
    
    for field, option in layout.iter_options():
        cell_coord = option.cell
        assert cell_coord in cells, (
            f"Field {field.key!r} option {option.code!r} references missing cell {cell_coord}"
        )
        
        cell_text = cells[cell_coord]
        normalized_cell = cell_text.replace(" ", "").replace("\n", "")
        normalized_label = option.label.replace(" ", "").replace("\n", "")
        
        # Check that the label appears in the cell text
        # Handle special cases:
        # - Most cells: □<label>
        # - A48: label comes before checkbox
        # - B23: contains both identity label and extra text
        assert normalized_label in normalized_cell, (
            f"Field {field.key!r} option {option.code!r} at {cell_coord}: "
            f"label {option.label!r} not found in cell text {cell_text!r}"
        )
        
        # Verify checkbox character is present
        assert "□" in cell_text, (
            f"Field {field.key!r} option {option.code!r} at {cell_coord}: "
            f"cell does not contain checkbox character '□'"
        )


def test_every_sheet_checkbox_option_is_modeled():
    """Every checkbox cell in the workbook (except aggregate counts) must be modeled."""
    layout = service_record_layout()
    cells = _read_workbook_cells()
    
    # Collect all checkbox cells from the workbook
    checkbox_cells = {coord for coord, text in cells.items() if "□" in text}
    
    # Remove non-option aggregate count cells
    checkbox_cells = checkbox_cells - _NON_OPTION_CHECKBOX_CELLS
    
    # Collect all modeled option cells
    modeled_cells = {opt.cell for _, opt in layout.iter_options()}
    
    # Check for any missing cells
    missing = checkbox_cells - modeled_cells
    assert not missing, (
        f"Workbook has {len(missing)} checkbox cells not in model: {sorted(missing)}"
    )
    
    # Also check that we're not modeling cells that don't exist
    extra = modeled_cells - checkbox_cells
    assert not extra, (
        f"Model has {len(extra)} option cells not in workbook checkbox cells: {sorted(extra)}"
    )


def test_every_record_path_is_legal():
    """Every non-None Field.record_path must be a legal path into domain.Record."""
    import dataclasses
    from ocr_from2xlsx import constants, domain
    
    layout = service_record_layout()
    
    # Get field names for each domain dataclass
    record_fields = {f.name for f in dataclasses.fields(domain.Record)}
    patient_fields = {f.name for f in dataclasses.fields(domain.PatientFields)}
    services_fields = {f.name for f in dataclasses.fields(domain.Services)}
    
    for field in layout.iter_fields():
        path = field.record_path
        if path is None:
            # diagnosis_date has record_path=None and should be excluded
            continue
        
        parts = path.split(".")
        
        # Top-level Record field
        if len(parts) == 1:
            assert parts[0] in record_fields, (
                f"Field {field.key!r} record_path {path!r}: "
                f"unknown Record field {parts[0]!r}"
            )
        
        # patient_fields.* paths
        elif len(parts) == 2 and parts[0] == "patient_fields":
            assert parts[1] in patient_fields, (
                f"Field {field.key!r} record_path {path!r}: "
                f"unknown PatientFields field {parts[1]!r}"
            )
        
        # services.* paths (not consultation subcategories)
        elif len(parts) == 2 and parts[0] == "services":
            assert parts[1] in services_fields, (
                f"Field {field.key!r} record_path {path!r}: "
                f"unknown Services field {parts[1]!r}"
            )
        
        # services.consultation.<category> paths
        elif len(parts) == 3 and parts[0] == "services" and parts[1] == "consultation":
            category = parts[2]
            assert category in constants.SERVICE_CATEGORIES, (
                f"Field {field.key!r} record_path {path!r}: "
                f"unknown consultation category {category!r}"
            )
        
        else:
            raise AssertionError(
                f"Field {field.key!r} record_path {path!r}: "
                f"unexpected path shape (expected Record.field, patient_fields.field, "
                f"services.field, or services.consultation.category)"
            )
