from __future__ import annotations

from pathlib import Path

import pytest

from openpyxl import load_workbook

from ocr_from2xlsx.session import ImportSession
from ocr_from2xlsx.constants import BASIC_COLUMN_BY_FIELD
from tests.fixtures import create_workbook_template
from tests.test_json_io import make_record


def _column_for_header(sheet, header: str) -> int:
    for cell in sheet[1]:
        if cell.value == header:
            return cell.column
    raise AssertionError(f"Missing header in fixture: {header}")


def test_auto_confirm_writes_and_saves_record(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    session = ImportSession.start(template, working)
    result = session.accept_scan(make_record())
    session.close()

    assert result.status == "written"
    assert result.row_number == 2
    assert working.exists()
    wb = load_workbook(working)
    ws = wb["個案總表"]
    name_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["name"])
    id_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["medical_record_no"])
    assert ws.cell(row=2, column=name_col).value == "王小明"
    assert ws.cell(row=2, column=id_col).value == "A123456"
    wb.close()


def test_blocked_record_is_not_written(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    session = ImportSession.start(template, working)
    record = make_record()
    record.service_date = ""
    result = session.accept_scan(record)
    session.close()

    assert result.status == "blocked"
    assert result.row_number is None
    assert "service_date.invalid" in result.blockers
    wb = load_workbook(working)
    ws = wb["個案總表"]
    name_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["name"])
    id_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["medical_record_no"])
    assert ws.cell(row=2, column=name_col).value is None
    assert ws.cell(row=2, column=id_col).value is None
    wb.close()


def test_blocked_record_does_not_reserve_duplicate_key(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    session = ImportSession.start(template, working)
    blocked_record = make_record("scan-0001")
    blocked_record.patient_fields.nationality = None
    blocked_result = session.accept_scan(blocked_record)
    valid_record = make_record("scan-0002")
    valid_result = session.accept_scan(valid_record)
    session.close()

    assert blocked_result.status == "blocked"
    assert "patient.nationality.required" in blocked_result.blockers
    assert valid_result.status == "written"
    assert valid_result.row_number == 2
    assert "duplicate.in_batch" not in valid_result.blockers


def test_force_with_non_writable_blocker_is_blocked(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    session = ImportSession.start(template, working)
    record = make_record()
    record.service_date = "not-a-date"
    result = session.accept_scan(record, force=True)
    session.close()

    assert result.status == "blocked"
    assert result.row_number is None
    assert "service_date.invalid" in result.blockers
    assert "force.non_writable" in result.blockers
    wb = load_workbook(working)
    ws = wb["個案總表"]
    name_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["name"])
    id_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["medical_record_no"])
    assert ws.cell(row=2, column=name_col).value is None
    assert ws.cell(row=2, column=id_col).value is None
    wb.close()


def test_force_with_patient_blocker_writes_record(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    session = ImportSession.start(template, working)
    record = make_record()
    record.patient_fields.nationality = None
    result = session.accept_scan(record, force=True)
    session.close()

    assert result.status == "forced"
    assert result.row_number == 2
    assert "patient.nationality.required" in result.blockers
    wb = load_workbook(working)
    ws = wb["個案總表"]
    name_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["name"])
    id_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["medical_record_no"])
    assert ws.cell(row=2, column=name_col).value == "王小明"
    assert ws.cell(row=2, column=id_col).value == "A123456"
    wb.close()


def test_unconfirmed_name_blocks_even_when_forced_without_human_confirmation(
    tmp_path: Path,
) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    session = ImportSession.start(template, working)
    record = make_record()
    record.ocr.warnings = ["name.unconfirmed"]
    result = session.accept_scan(record, force=True)
    session.close()

    assert result.status == "blocked"
    assert result.row_number is None
    assert "name.unconfirmed" in result.blockers
    assert "force.non_writable" in result.blockers
    wb = load_workbook(working)
    ws = wb["個案總表"]
    name_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["name"])
    id_col = _column_for_header(ws, BASIC_COLUMN_BY_FIELD["medical_record_no"])
    assert ws.cell(row=2, column=name_col).value is None
    assert ws.cell(row=2, column=id_col).value is None
    wb.close()


def test_unconfirmed_name_blocks_by_default(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    record = make_record()
    record.identity = "public_other"
    record.patient_fields = None  # type: ignore[assignment]
    from ocr_from2xlsx.domain import PatientFields
    record.patient_fields = PatientFields()
    record.ocr.warnings = ["name.unconfirmed"]

    session = ImportSession.start(template, working)
    result = session.accept_scan(record)
    session.close()

    assert result.status == "blocked"
    assert "name.unconfirmed" in result.blockers


def test_allow_unconfirmed_name_writes_and_keeps_warning(tmp_path: Path) -> None:
    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)

    record = make_record()
    record.identity = "public_other"
    from ocr_from2xlsx.domain import PatientFields
    record.patient_fields = PatientFields()
    record.ocr.warnings = ["name.unconfirmed"]

    session = ImportSession.start(template, working)
    result = session.accept_scan(record, allow_unconfirmed_name=True)
    session.close()

    assert result.status in ("written", "forced")
    assert result.row_number is not None
    assert "name.unconfirmed" in record.ocr.warnings        # NOT stripped
    assert "name.unconfirmed" in result.warnings


def test_writer_failure_does_not_reserve_duplicate_key() -> None:
    class FailingWriter:
        def existing_duplicate_keys(self) -> set[tuple[str, str, str, str]]:
            return set()

        def write_record(self, record) -> int:
            return 2

        def save(self) -> None:
            raise RuntimeError("save failed")

        def close(self) -> None:
            return None

    session = ImportSession(FailingWriter())
    record = make_record()

    with pytest.raises(RuntimeError, match="save failed"):
        session.accept_scan(record)

    assert record.duplicate_key() not in session.batch_duplicate_keys


def test_accept_scan_overwrite_row_writes_to_that_row(tmp_path: Path) -> None:
    from ocr_from2xlsx.constants import GENDER_LABELS

    template = tmp_path / "template.xlsx"
    working = tmp_path / "working.xlsx"
    create_workbook_template(template)
    session = ImportSession.start(template, working)
    try:
        first = make_record("a")
        first.name = "王小明"
        first.medical_record_no = "A1"
        second = make_record("b")
        second.name = "李大華"
        second.medical_record_no = "B2"
        r1 = session.accept_scan(first, human_confirmed=True).row_number
        session.accept_scan(second, human_confirmed=True)

        # Re-open the first record (same duplicate key) and overwrite its row with a fix.
        corrected = make_record("a2")
        corrected.name = "王小明"
        corrected.medical_record_no = "A1"
        corrected.gender = "male"
        result = session.accept_scan(corrected, human_confirmed=True, overwrite_row=r1)

        assert result.row_number == r1
        assert result.status in {"written", "forced"}
        # Re-using the same key on an overwrite must not be flagged as an in-batch dup.
        assert "duplicate.in_batch" not in result.blockers
    finally:
        session.close()

    wb = load_workbook(working)
    try:
        sheet = wb["個案總表"]
        gender_col = _column_for_header(sheet, BASIC_COLUMN_BY_FIELD["gender"])
        assert sheet.cell(row=r1, column=gender_col).value == GENDER_LABELS["male"]
    finally:
        wb.close()
