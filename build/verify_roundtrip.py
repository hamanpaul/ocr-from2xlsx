"""Auditable end-to-end self-check for the GUI write path (#service-write-mapping,
#xlsx-image-save). Run this before handing off ANY change that touches the form / record /
workbook-write code:

    PYTHONPATH=src python build/verify_roundtrip.py

What it does (the same shape the operator drives, but deterministic):
  1. GENERATE a golden JSON batch with known selections (incl. the tricky cases that bit us:
     fatigue_strength, the no-period 院外 option, option 10, every 5-wide group, a non-patient
     record, and a name-only record with blank optionals).
  2. LOAD it the way the GUI does (JsonRecordSource) and push each record through the REAL
     ConfirmForm widgets (prefill -> collect -> apply_form_state) — i.e. "GUI 讀入 -> GUI 寫入".
  3. WRITE to a real .xlsx via ImportSession.accept_scan, against the REAL official template if
     present (so embedded-image saves are exercised), else a generated full+image template.
  4. READ the .xlsx back and ASSERT it matches the golden JSON cell-by-cell: every selection in
     its own numbered column with the real label, NO raw English code, file opens (not corrupt),
     images preserved.

Exit 0 + "PASS" when everything round-trips; non-zero + a per-record diff otherwise. The point
is that *this script* — not the human — is the QA: reproducible and auditable.
"""
from __future__ import annotations

import sys
import tempfile
from pathlib import Path

REPO = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO / "src"))

from ocr_from2xlsx.capture import JsonRecordSource  # noqa: E402
from ocr_from2xlsx.confirm_form import apply_form_state, record_to_form_state  # noqa: E402
from ocr_from2xlsx.constants import (  # noqa: E402
    AGE_GROUP_LABELS,
    BASIC_COLUMN_BY_FIELD,
    CANCER_LABELS,
    CHANNEL_LABELS,
    DISEASE_STATUS_LABELS,
    GENDER_LABELS,
    IDENTITY_LABELS,
    NATIONALITY_LABELS,
    SOURCE_LABELS,
)
from ocr_from2xlsx.domain import (  # noqa: E402
    Batch,
    OcrInfo,
    PatientFields,
    Record,
    ReviewInfo,
    Services,
    SourceBatch,
)
from ocr_from2xlsx.form_layout import service_record_layout  # noqa: E402
from ocr_from2xlsx.json_io import dump_batch  # noqa: E402
from ocr_from2xlsx.session import ImportSession  # noqa: E402

REAL_TEMPLATE = REPO / "115年整年月報表統計_單案統計加總版(下拉式)(空白).xlsx"

SERVICE_PREFIX = {
    "services.consultation.health_medical": "諮詢-健康與醫療系統",
    "services.consultation.symptom_side_effect": "諮詢-症狀與副作用照護",
    "services.consultation.nutrition_diet": "諮詢-營養與飲食",
    "services.consultation.psychosocial_emotion": "諮詢-社會心理情緒",
    "services.consultation.financial_social": "諮詢-經濟與社會資源",
    "services.consultation.care_support": "諮詢-照顧與支持",
    "services.supplies": "提供實體用品及設備",
    "services.internal_referrals": "轉介或連結院內資源",
    "services.external_referrals": "轉介或連結院外資源",
    "services.referral_outcomes": "轉介或連結資源成果",
}


def _golden_records() -> list[Record]:
    """Deterministic records that cover the cases that have actually broken before."""
    patient = Record(
        record_id="golden-patient",
        service_date="2026-03-15",
        identity="patient",
        name="王小明",
        medical_record_no="A123456",
        gender="female",
        patient_fields=PatientFields(
            nationality="local",
            age_group="51_60",
            channel="internal_referral",
            disease_status="treating",
            source="outpatient",
            cancers=["breast_cancer", "lung_cancer", "liver_cancer"],
            newly_diagnosed_within_year=True,
        ),
        services=Services(
            consultation={
                "health_medical": ["screening_prevention", "other"],
                # fatigue_strength = symptom option 4 -> must land in 諮詢-症狀與副作用照護4
                "symptom_side_effect": ["fatigue_strength", "wound_care"],
                "financial_social": ["rehab_supplies_aids", "other"],  # options 7 & 8
            },
            supplies=["wig_hat", "other_equipment"],
            # internal social_welfare = option 3, psychology = option 5
            internal_referrals=["social_welfare", "psychology"],
            # external wig_hat = option 1 (NO period in label), other_activity = option 10
            external_referrals=["wig_hat", "other_activity"],
            referral_outcomes=["received_service_help"],  # option 4
        ),
        discharge_followup=True,
        ocr=OcrInfo(confidence=0.95, raw_text="raw", warnings=[]),
        review=ReviewInfo(status="pending", edited_by_user=False),
    )
    family = Record(
        record_id="golden-family",
        service_date="2026-03-16",
        identity="family_caregiver",
        name="李大華",
        medical_record_no="B654321",
        gender="male",
        services=Services(
            consultation={"care_support": ["caregiver_support", "discharge_planning"]},  # 3 & 5
            referral_outcomes=["received_wig_hat"],  # option 1
        ),
        ocr=OcrInfo(confidence=0.9, raw_text="raw", warnings=[]),
        review=ReviewInfo(status="pending", edited_by_user=False),
    )
    name_only = Record(
        record_id="golden-name-only",
        service_date="",  # blank optional fields — relaxed write must still succeed
        identity="",
        name="陳小美",
        medical_record_no="",
        gender="",
        ocr=OcrInfo(confidence=None, raw_text="raw", warnings=[]),
        review=ReviewInfo(status="pending", edited_by_user=False),
    )
    return [patient, family, name_only]


def _expected_cancers(record: Record) -> set[str]:
    """The 癌別1/2/3 slots are interchangeable and the GUI form stores multi-choice values
    sorted by code, so compare them as a set of labels (order-independent), not per-slot."""
    if record.identity != "patient":
        return set()
    return {
        CANCER_LABELS.get(c, c) for c in sorted(record.patient_fields.cancers)[:3]
    }


def _expected_cells(record: Record, layout) -> dict[str, object]:
    """The cells the XLSX SHOULD contain for this record, derived independently from
    form_layout + constants (option position = column number; option label = cell value).
    癌別 slots are checked separately as a set (see _expected_cancers)."""
    cells: dict[str, object] = {}
    if record.service_date:
        try:
            cells[BASIC_COLUMN_BY_FIELD["service_month"]] = record.service_month_label()
        except ValueError:
            pass
        cells[BASIC_COLUMN_BY_FIELD["service_date"]] = record.service_date
    if record.identity:
        cells[BASIC_COLUMN_BY_FIELD["identity"]] = IDENTITY_LABELS[record.identity]
    if record.name:
        cells[BASIC_COLUMN_BY_FIELD["name"]] = record.name
    if record.medical_record_no:
        cells[BASIC_COLUMN_BY_FIELD["medical_record_no"]] = record.medical_record_no
    if record.gender:
        cells[BASIC_COLUMN_BY_FIELD["gender"]] = GENDER_LABELS[record.gender]
    if record.discharge_followup is not None:
        cells[BASIC_COLUMN_BY_FIELD["discharge_followup"]] = "是" if record.discharge_followup else "否"
    if record.identity == "patient":
        pf = record.patient_fields
        if pf.nationality:
            cells[BASIC_COLUMN_BY_FIELD["nationality"]] = NATIONALITY_LABELS[pf.nationality]
        if pf.age_group:
            cells[BASIC_COLUMN_BY_FIELD["age_group"]] = AGE_GROUP_LABELS[pf.age_group]
        if pf.channel:
            cells[BASIC_COLUMN_BY_FIELD["channel"]] = CHANNEL_LABELS[pf.channel]
        if pf.disease_status:
            cells[BASIC_COLUMN_BY_FIELD["disease_status"]] = DISEASE_STATUS_LABELS[pf.disease_status]
        if pf.source:
            cells[BASIC_COLUMN_BY_FIELD["source"]] = SOURCE_LABELS[pf.source]
        if pf.newly_diagnosed_within_year is not None:
            cells[BASIC_COLUMN_BY_FIELD["newly_diagnosed_within_year"]] = (
                "是" if pf.newly_diagnosed_within_year else "否"
            )

    groups: dict[str, list[str]] = {
        f"services.consultation.{cat}": list(codes)
        for cat, codes in record.services.consultation.items()
    }
    groups["services.supplies"] = list(record.services.supplies)
    groups["services.internal_referrals"] = list(record.services.internal_referrals)
    groups["services.external_referrals"] = list(record.services.external_referrals)
    groups["services.referral_outcomes"] = list(record.services.referral_outcomes)
    for path, codes in groups.items():
        prefix = SERVICE_PREFIX[path]
        field = layout.field_by_key(path[len("services."):])
        by_code = {opt.code: (i, opt.label) for i, opt in enumerate(field.options, start=1)}
        for code in codes:
            number, label = by_code[code]
            cells[f"{prefix}{number}"] = label
    return cells


def _build_template_with_image(path: Path) -> None:
    """Fallback template (used when the real one is absent): the full 個案總表 column set plus an
    embedded image, so the round-trip still exercises image-bearing saves."""
    from io import BytesIO

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage

    groups = [
        ("諮詢-健康與醫療系統", 8), ("諮詢-症狀與副作用照護", 7), ("諮詢-營養與飲食", 4),
        ("諮詢-社會心理情緒", 6), ("諮詢-經濟與社會資源", 8), ("諮詢-照顧與支持", 6),
        ("提供實體用品及設備", 4), ("轉介或連結院內資源", 10), ("轉介或連結院外資源", 10),
        ("轉介或連結資源成果", 4),
    ]
    headers = [
        "序", "服務月份", "身分", "服務日期", "姓名", "ID", "生日", "性別", "是否曾經今年服務過",
        "國籍\n(病人才填)", "年齡\n(病人才填)", "管道\n(病人才填)", "疾病狀態\n(病人才填)",
        "來源\n(病人才填)", "癌別1\n(病人才填)", "癌別2\n(病人才填)", "癌別3\n(病人才填)",
        "一年內新診斷(病人才填)", "出院後關懷",
    ]
    for prefix, count in groups:
        headers.extend(f"{prefix}{i}" for i in range(1, count + 1))
    wb = Workbook()
    ws = wb.active
    ws.title = "個案總表"
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)
    buf = BytesIO()
    PILImage.new("RGB", (8, 8), "red").save(buf, format="PNG")
    buf.seek(0)
    ws.add_image(XLImage(buf), "ZZ1")
    wb.save(path)


def main() -> int:
    try:
        import tkinter as tk
    except Exception as exc:  # pragma: no cover
        print(f"SKIP: tkinter unavailable ({exc})")
        return 0
    try:
        root = tk.Tk()
    except tk.TclError as exc:
        print(f"SKIP: no display for Tk ({exc})")
        return 0
    root.withdraw()

    from ocr_from2xlsx.app import ConfirmForm

    layout = service_record_layout()
    workdir = Path(tempfile.mkdtemp(prefix="verify_roundtrip_"))
    golden_json = workdir / "golden.json"
    working = workdir / "匯入中.xlsx"

    records = _golden_records()
    dump_batch(
        Batch(
            source_batch=SourceBatch(
                created_at="2026-01-01T00:00:00+08:00",
                source_type="json_import",
                template_name="verify",
            ),
            records=records,
        ),
        golden_json,
    )

    if REAL_TEMPLATE.is_file():
        template = REAL_TEMPLATE
        template_note = f"real template ({template.name})"
        expect_images = True
    else:
        template = workdir / "template.xlsx"
        _build_template_with_image(template)
        template_note = "generated full+image fallback template"
        expect_images = True

    loaded = list(JsonRecordSource(golden_json).records())
    form = ConfirmForm(root, layout)
    session = ImportSession.start(template, working)
    write_errors: list[str] = []
    for record in loaded:
        # GUI 讀入 (prefill widgets) -> GUI 寫入 (collect widgets back -> apply -> write)
        form.prefill(record_to_form_state(layout, record))
        apply_form_state(layout, record, form.collect())
        try:
            result = session.accept_scan(record, human_confirmed=True, relaxed=True)
            if result.status not in {"written", "forced"}:
                write_errors.append(f"{record.record_id}: status={result.status} blockers={result.blockers}")
        except Exception as exc:  # the image-save crash surfaced here
            write_errors.append(f"{record.record_id}: write raised {type(exc).__name__}: {exc}")
    session.close()
    root.destroy()

    # READ BACK + ASSERT vs the golden JSON
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    failures: list[str] = []
    failures.extend(write_errors)
    try:
        wb = load_workbook(working, data_only=True)
    except Exception as exc:
        print(f"FAIL: output workbook is corrupt / unreadable: {type(exc).__name__}: {exc}")
        return 1
    ws = wb["個案總表"]
    header_to_col = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    all_service_codes = {opt.code for _f, opt in layout.iter_options()}
    for row_offset, record in enumerate(records):
        row = 2 + row_offset
        expected = _expected_cells(record, layout)
        for header, want in expected.items():
            col = header_to_col.get(header)
            if col is None:
                failures.append(f"[{record.record_id}] missing column {header!r} in template")
                continue
            got = ws.cell(row=row, column=col).value
            if str(got) != str(want):
                failures.append(
                    f"[{record.record_id}] {get_column_letter(col)}{row} ({header!r}): "
                    f"expected {want!r}, got {got!r}"
                )
        # 癌別 slots compared as a set (order-independent)
        want_cancers = _expected_cancers(record)
        got_cancers = {
            ws.cell(row=row, column=header_to_col[h]).value
            for h in ("癌別1\n(病人才填)", "癌別2\n(病人才填)", "癌別3\n(病人才填)")
            if h in header_to_col and ws.cell(row=row, column=header_to_col[h]).value not in (None, "")
        }
        if got_cancers != want_cancers:
            failures.append(
                f"[{record.record_id}] 癌別 set mismatch: expected {sorted(want_cancers)}, got {sorted(got_cancers)}"
            )
        # no raw English service code leaked anywhere in the row
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=row, column=col).value
            if isinstance(val, str) and val in all_service_codes:
                failures.append(
                    f"[{record.record_id}] {get_column_letter(col)}{row}: raw code {val!r} leaked"
                )

    image_count = sum(len(getattr(sh, "_images", [])) for sh in load_workbook(working).worksheets)
    wb.close()
    if expect_images and image_count == 0:
        failures.append("embedded images were dropped (expected the template's logo to survive)")

    print(f"template: {template_note}")
    print(f"records: {len(records)} | output: {working}")
    print(f"images preserved: {image_count}")
    if failures:
        print(f"\nFAIL — {len(failures)} problem(s):")
        for f in failures:
            print(f"  - {f}")
        return 1
    print("\nPASS — every golden selection round-tripped to the correct cell, no code leaks, "
          "file valid, images preserved.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
